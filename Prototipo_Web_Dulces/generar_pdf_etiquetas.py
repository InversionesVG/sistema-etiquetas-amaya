#!/usr/bin/env python3
"""
Generador de PDFs - AMAYA EXPRESS
VERSIÓN CORREGIDA - ETIQUETAS LÁCTEAS
"""
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor, black
import sys, os, re
import datetime

LABEL_CONFIGS = {
    'AVERY_8164': {
        'width': 3.33 * inch, 'height': 4 * inch, 'columns': 2, 'rows': 3, 'per_page': 6,
        'margin_left': 0.16 * inch, 'margin_top': 0.48 * inch, 
        'h_spacing': 0.95 * inch, 'v_spacing': -0.68 * inch,
    },
    'LACTEO_AVERY_8164': {
        # ETIQUETA HORIZONTAL (girada 90°)
        'width': 4 * inch,      
        'height': 3.33 * inch,  
        'columns': 2, 
        'rows': 3, 
        'per_page': 6,
        'margin_left': 0.16 * inch,   
        'margin_top': 0.48 * inch,    
        'h_spacing': 0.18 * inch,     # Espacio horizontal entre etiquetas
        'v_spacing': 0.05 * inch,     # AJUSTADO: Reducido de 0.15" a 0.05"
    },
    'PLS504': {
        'width': 3 * inch, 'height': 2 * inch, 'columns': 2, 'rows': 5, 'per_page': 10,
        'margin_left': 0.25 * inch, 'margin_top': 0.5 * inch,
        'h_spacing': 0.25 * inch, 'v_spacing': 0 * inch,
    },
    'AVERY_5260': {
        # Avery 5260: 1" x 2-5/8", 3 columnas x 10 filas = 30 por hoja (specs oficiales Avery)
        'width': 2.625 * inch, 'height': 1 * inch, 'columns': 3, 'rows': 10, 'per_page': 30,
        'margin_left': 0.1875 * inch, 'margin_top': 0.5 * inch,
        'h_spacing': 0.125 * inch, 'v_spacing': 0 * inch,
    }
}

AMAYA_BLUE = HexColor('#0066B3')
DARK_BLUE = HexColor('#003366')  # Para cuadro de peso
LOGO_PATH = "Amaya_Imagen.png"
FONDO_SUPERIOR_PATH = "Fondo_Superior.png"
FONDO_INFERIOR_PATH = "Fondo_Inferior.png"
PAGE_WIDTH, PAGE_HEIGHT = letter
PADDING = 0.1 * inch

FDA_REF = {
    'Total_Fat': 78, 'Saturated_Fat': 20, 'Cholesterol': 300, 'Sodium': 2300,
    'Total_Carbohydrate': 275, 'Dietary_Fiber': 28, 'Added_Sugars': 50, 
    'Protein': 50, 'Vitamin_D': 20, 'Calcium': 1300, 'Iron': 18, 'Potassium': 4700
}

def extraer_numero(t):
    m = re.search(r'(\d+\.?\d*)', str(t).strip()) if t and str(t).strip() else None
    return float(m.group(1)) if m else 0

def calc_dv(n, v):
    val = extraer_numero(v)
    return round((val / FDA_REF[n]) * 100) if n in FDA_REF and val > 0 else None

def es_cero(valor):
    """
    True si el valor está vacío/None, o si el primer número que contiene es 0.
    Se usa en AVERY_5260 para omitir filas de Nutrition Facts en 0 (p.ej. "0g", "0mg")
    y así aprovechar el espacio para lo que sí tiene valor.
    NOTA: "<1g" NO se considera cero (el número detectado es 1), así que sí se muestra.
    """
    if valor is None:
        return True
    s = str(valor).strip()
    if not s or s == 'None':
        return True
    m = re.search(r'(\d+\.?\d*)', s)
    if not m:
        return True
    return float(m.group(1)) == 0

def formatear_exp(valor):
    """
    Formatea Expiration_Date como MM/AAAA (igual que en la etiqueta de muestra:
    "EXP:12/2027"). Si Excel guardó la celda como fecha, evita que salga con la
    hora pegada (ej. "2027-04-14 00:00:00"). Si ya viene como texto (ej. algún
    producto donde se escribió "26/07/2026" a mano), se deja tal cual.
    """
    if isinstance(valor, (datetime.datetime, datetime.date)):
        return valor.strftime('%m/%Y')
    return str(valor) if valor not in (None, '') else ''

def dibujar_lacteo_avery(c, x, y, p, cfg):
    """
    VERSIÓN CORREGIDA - Etiqueta láctea horizontal
    
    Layout corregido:
    ┌──────────────────────────────────────────────┐
    │ FONDO AMARILLO (60% superior)                │
    │ KEEP REFRIGERATED (arriba derecha)           │
    │ Product_Name (grande, centrado)              │
    │ Product_Name_English (cursiva, centrado)     │
    │ (espacio para vacas del fondo)               │
    ├──────────────────────────────────────────────┤
    │ FONDO AZUL (40% inferior)                    │
    │ INGREDIENTS: ... (arriba izquierda)          │
    │ CONTAINS: MILK PROTEINS                      │
    │                                              │
    │ [LOGO]  [PESO]  Distributed/  ┌───────────┐ │
    │ AMAYA   azul    Imported      │ Nutrition │ │
    │                               │ Facts     │ │
    │ (908)405-5553  email  (908).. └───────────┘ │
    └──────────────────────────────────────────────┘
    """
    
    # CORRECCIÓN: División 60/40 en lugar de 50/50
    altura_superior = cfg['height'] * 0.6
    altura_inferior = cfg['height'] * 0.4
    
    # ========================================================================
    # FONDOS
    # ========================================================================
    if os.path.exists(FONDO_SUPERIOR_PATH):
        c.drawImage(
            FONDO_SUPERIOR_PATH, 
            x, 
            y + altura_inferior,
            width=cfg['width'], 
            height=altura_superior,
            preserveAspectRatio=False,
            mask='auto'
        )
    
    if os.path.exists(FONDO_INFERIOR_PATH):
        c.drawImage(
            FONDO_INFERIOR_PATH, 
            x, 
            y,
            width=cfg['width'], 
            height=altura_inferior,
            preserveAspectRatio=False,
            mask='auto'
        )
    
    # ========================================================================
    # SECCIÓN SUPERIOR (60%) - SOBRE FONDO AMARILLO
    # ========================================================================
    c.setFillColor(DARK_BLUE)
    
    # KEEP REFRIGERATED (arriba derecha, más abajo)
    yp = y + cfg['height'] - PADDING * 0.3
    yp -= 8  # Bajar KEEP REFRIGERATED una línea
    c.setFont("Helvetica-Bold", 8)
    keep_text = "KEEP REFRIGERATED"
    text_width = c.stringWidth(keep_text, "Helvetica-Bold", 8)
    c.drawString(x + cfg['width'] - text_width - PADDING * 0.3, yp, keep_text)
    yp -= 15  # Espacio reducido hasta Product_Name
    
    # Product_Name - MÁS GRANDE, centrado
    nombre_producto = p.get('Product_Name', '')
    c.setFont("Helvetica-Bold", 20)  # Aumentado de 16 a 20
    text_width = c.stringWidth(nombre_producto, "Helvetica-Bold", 20)
    c.drawString(x + (cfg['width'] - text_width) / 2, yp, nombre_producto)
    yp -= 16
    
    # Product_Name_English - MÁS GRANDE, cursiva, centrado
    nombre_eng = p.get('Product_Name_English', '')
    if nombre_eng and str(nombre_eng).strip() != 'None':
        c.setFont("Helvetica-BoldOblique", 13)  # Aumentado de 10 a 13
        text_width = c.stringWidth(nombre_eng, "Helvetica-BoldOblique", 13)
        c.drawString(x + (cfg['width'] - text_width) / 2, yp, nombre_eng)
    
    # ========================================================================
    # SECCIÓN INFERIOR (40%) - SOBRE FONDO AZUL
    # ========================================================================
    
    # CORRECCIÓN 3 y 5: INGREDIENTS en la parte superior del fondo azul
    y_inicio_azul = y + altura_inferior - PADDING * 0.3
    yp_azul = y_inicio_azul
    
    c.setFillColor(black)
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x + PADDING * 0.3, yp_azul, "INGREDIENTS:")
    yp_azul -= 6
    
    # Ingredientes con ajuste de línea
    c.setFont("Helvetica", 5)
    ingredientes = str(p.get('Ingredients', ''))
    # CORRECCIÓN 4: Ancho máximo reducido para dejar espacio a Nutrition Facts
    max_width_ing = cfg['width'] - 1.3 * inch  # Deja espacio para NF
    linea = ""
    
    for palabra in ingredientes.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 5) < max_width_ing:
            linea = test
        else:
            c.drawString(x + PADDING * 0.3, yp_azul, linea.strip())
            yp_azul -= 5
            linea = palabra + " "
    
    if linea:
        c.drawString(x + PADDING * 0.3, yp_azul, linea.strip())
        yp_azul -= 7
    
    # CONTAINS: MILK PROTEINS
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x + PADDING * 0.3, yp_azul, "CONTAINS: MILK PROTEINS.")
    yp_azul -= 8
    
    # FECHA DE VENCIMIENTO debajo de CONTAINS
    c.setFont("Helvetica-Bold", 6)
    exp_text = f"EXP: {p.get('Expiration_Date', '')}"
    c.drawString(x + PADDING * 0.3, yp_azul, exp_text)
    yp_azul -= 10
    
    # CORRECCIÓN 5: LOGO AMAYA debajo de ingredientes
    if os.path.exists(LOGO_PATH):
        logo_size = 0.35 * inch
        c.drawImage(
            LOGO_PATH, 
            x + PADDING * 0.3, 
            yp_azul - logo_size,
            logo_size, 
            logo_size, 
            preserveAspectRatio=True, 
            mask='auto'
        )
    
    # CORRECCIÓN 6: NET WEIGHT con fondo AZUL OSCURO (no negro)
    c.setFillColor(DARK_BLUE)  # Azul oscuro #003366
    peso_x = x + PADDING * 0.3 + logo_size + 0.08 * inch
    peso_y = yp_azul - 0.35 * inch
    peso_ancho = 0.45 * inch
    peso_alto = 0.32 * inch
    
    c.rect(peso_x, peso_y, peso_ancho, peso_alto, fill=1, stroke=0)
    
    # Texto del peso en blanco
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 8)
    net_weight = str(p.get('Net_Weight', ''))
    
    if 'oz' in net_weight.lower():
        partes = net_weight.split('(')
        oz_text = partes[0].strip()
        gram_text = f"({partes[1]}" if len(partes) > 1 else ""
        
        tw = c.stringWidth(oz_text, "Helvetica-Bold", 8)
        c.drawString(peso_x + (peso_ancho - tw) / 2, peso_y + peso_alto - 11, oz_text)
        
        if gram_text:
            c.setFont("Helvetica", 5.5)
            tw = c.stringWidth(gram_text, "Helvetica", 5.5)
            c.drawString(peso_x + (peso_ancho - tw) / 2, peso_y + peso_alto - 20, gram_text)
    else:
        tw = c.stringWidth(net_weight, "Helvetica-Bold", 7)
        c.drawString(peso_x + (peso_ancho - tw) / 2, peso_y + peso_alto / 2 - 2, net_weight)
    
    # CORRECCIÓN 7: DISTRIBUTED BY / IMPORTED BY a la izquierda de NF
    c.setFillColor(DARK_BLUE)
    dist_x = peso_x + peso_ancho + 0.08 * inch
    dist_y = yp_azul - 0.05 * inch
    
    imported_by = p.get('Imported_By', '')
    has_imported = imported_by and str(imported_by).strip() and str(imported_by).strip() != 'None'
    
    c.setFont("Helvetica-Bold", 5)
    c.drawString(dist_x, dist_y, "Distributed by:")
    c.setFont("Helvetica", 4.5)
    c.drawString(dist_x, dist_y - 5.5, "Amaya Express International")
    c.drawString(dist_x, dist_y - 10, "472 Somerset St.")
    c.drawString(dist_x, dist_y - 14.5, "North Plainfield, NJ 06070")
    
    if has_imported:
        c.setFont("Helvetica-Bold", 5)
        c.drawString(dist_x, dist_y - 21, "Imported by:")
        
        imported_parts = str(imported_by).split(',')
        c.setFont("Helvetica", 4.5)
        if len(imported_parts) >= 1:
            c.drawString(dist_x, dist_y - 26.5, imported_parts[0].strip())
        if len(imported_parts) >= 2:
            c.drawString(dist_x, dist_y - 31, imported_parts[1].strip())
        if len(imported_parts) >= 3:
            c.drawString(dist_x, dist_y - 35.5, imported_parts[2].strip())
    
    # INFORMACIÓN DE CONTACTO (abajo)
    contact_y = y + PADDING * 0.25
    c.setFont("Helvetica", 4.5)
    c.drawString(x + PADDING * 0.3, contact_y, "(908) 405-5553")
    c.drawString(x + PADDING * 0.3 + 0.6 * inch, contact_y, "amayaexpress21@hotmail.com")
    c.drawString(x + PADDING * 0.3 + 1.65 * inch, contact_y, "(908) 405-3072")
    
    # ========================================================================
    # CORRECCIÓN 4: NUTRITION FACTS (más estrecho, lado derecho)
    # ========================================================================
    nf_width = 1.05 * inch  # Reducido de 1.15" a 1.05"
    nf_x = x + cfg['width'] - nf_width - PADDING * 0.2
    nf_y_bottom = y + PADDING * 0.4
    nf_height = altura_inferior - PADDING * 0.7
    
    # Fondo blanco
    c.setFillColor(HexColor('#FFFFFF'))
    c.setStrokeColor(black)
    c.setLineWidth(1)
    c.rect(nf_x, nf_y_bottom, nf_width, nf_height, fill=1, stroke=1)
    
    # Contenido
    c.setFillColor(black)
    yp_nf = nf_y_bottom + nf_height - 5
    
    c.setFont("Helvetica-Bold", 7)
    c.drawString(nf_x + 3, yp_nf, "Nutrition Facts")
    yp_nf -= 7
    
    c.setFont("Helvetica", 4.5)
    servings = p.get('Servings_Per_Container', '')
    c.drawString(nf_x + 3, yp_nf, f"{servings} servings per Container")
    yp_nf -= 5.5
    
    c.setFont("Helvetica-Bold", 4.5)
    serving_size = p.get('Serving_Size', '')
    # Ajustar texto si es muy largo
    ss_text = f"Serving Size {serving_size}"
    if c.stringWidth(ss_text, "Helvetica-Bold", 4.5) > nf_width - 6:
        c.setFont("Helvetica-Bold", 4)
        ss_text = f"Size {serving_size}"
    c.drawString(nf_x + 3, yp_nf, ss_text)
    yp_nf -= 2
    
    c.setLineWidth(1.2)
    c.line(nf_x + 3, yp_nf, nf_x + nf_width - 3, yp_nf)
    yp_nf -= 4
    
    c.setFont("Helvetica", 3.5)
    c.drawString(nf_x + 3, yp_nf, "Amount per serving")
    yp_nf -= 5.5
    
    # CORREGIDO: Calories del mismo tamaño que otros nutrientes
    c.setFont("Helvetica-Bold", 3.8)  # Mismo tamaño que nutrientes
    c.drawString(nf_x + 3, yp_nf, "Calories")
    calories = p.get('Calories', '')
    c.drawString(nf_x + 28, yp_nf, str(calories))
    yp_nf -= 2
    
    c.setLineWidth(2)
    c.line(nf_x + 3, yp_nf, nf_x + nf_width - 3, yp_nf)
    yp_nf -= 4
    
    c.setFont("Helvetica-Bold", 3.5)
    c.drawRightString(nf_x + nf_width - 3, yp_nf, "% Daily Value*")
    yp_nf -= 5.5
    
    c.setLineWidth(0.3)
    
    def dn_nf(label, campo, bold=True, indent=0):
        nonlocal yp_nf
        valor = p.get(campo, '')
        dv = calc_dv(campo, valor)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 3.8)
        c.drawString(nf_x + 3 + indent, yp_nf, f"{label} {valor}")
        if dv:
            c.drawRightString(nf_x + nf_width - 3, yp_nf, f"{dv}%")
        yp_nf -= 1
        c.line(nf_x + 3, yp_nf, nf_x + nf_width - 3, yp_nf)
        yp_nf -= 4.5
    
    dn_nf("Total Fat", "Total_Fat")
    dn_nf("Sat. Fat", "Saturated_Fat", False, 4)
    dn_nf("Trans Fat", "Trans_Fat", False, 4)
    dn_nf("Cholesterol", "Cholesterol")
    dn_nf("Sodium", "Sodium")
    dn_nf("Total Carb.", "Total_Carbohydrate")
    dn_nf("Fiber", "Dietary_Fiber", False, 4)
    dn_nf("Sugars", "Total_Sugars", False, 4)
    dn_nf("  Added", "Added_Sugars", False, 6)
    
    c.setFont("Helvetica-Bold", 3.8)
    pv = p.get('Protein', '')
    pd = calc_dv('Protein', pv)
    c.drawString(nf_x + 3, yp_nf, f"Protein {pv}")
    if pd:
        c.drawRightString(nf_x + nf_width - 3, yp_nf, f"{pd}%")
    yp_nf -= 1
    
    # Línea final después de Protein
    c.setLineWidth(1.5)
    c.line(nf_x + 3, yp_nf, nf_x + nf_width - 3, yp_nf)
    # FIN de Nutrition Facts (sin micronutrientes ni leyenda)

def dibujar_avery(c, x, y, p, cfg):
    """Versión normal Avery 8164"""
    yp = y + cfg['height'] - PADDING
    c.setFillColor(AMAYA_BLUE)
    c.setStrokeColor(black)
    
    logo_size = 0.4 * inch
    if os.path.exists(LOGO_PATH):
        c.drawImage(LOGO_PATH, x + PADDING, yp - logo_size, 
                   logo_size, logo_size, preserveAspectRatio=True)
    
    nombre_esp = p.get('Product_Name', '')
    nombre_eng = p.get('Product_Name_English', '')
    
    nombre_x_start = x + PADDING + logo_size + 0.05 * inch
    nombre_width = cfg['width'] - logo_size - PADDING - 0.05 * inch
    
    c.setFont("Helvetica-Bold", 9)
    tw = c.stringWidth(nombre_esp, "Helvetica-Bold", 9)
    c.drawString(nombre_x_start + (nombre_width - tw) / 2, yp - 0.12 * inch, nombre_esp)
    
    if nombre_eng and str(nombre_eng).strip() != 'None':
        c.setFont("Helvetica", 7)
        tw = c.stringWidth(nombre_eng, "Helvetica", 7)
        c.drawString(nombre_x_start + (nombre_width - tw) / 2, yp - 0.24 * inch, nombre_eng)
    
    yp = yp - logo_size - 0.05 * inch
    c.setFont("Helvetica", 7)
    net_weight = f"Net Wt. {p.get('Net_Weight', '')}"
    tw = c.stringWidth(net_weight, "Helvetica", 7)
    c.drawString(nombre_x_start + (nombre_width - tw) / 2, yp, net_weight)
    yp -= 10
    
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x + PADDING, yp, "Nutrition Facts")
    yp -= 10
    
    c.setFont("Helvetica", 6)
    c.drawString(x + PADDING, yp, f"{p.get('Servings_Per_Container', '')} servings per container")
    yp -= 7
    
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x + PADDING, yp, f"Serving size    {p.get('Serving_Size', '')}")
    yp -= 2
    
    c.setLineWidth(1.5)
    c.line(x + PADDING, yp, x + cfg['width'] - PADDING, yp)
    yp -= 5
    
    c.setFont("Helvetica", 5)
    c.drawString(x + PADDING, yp, "Amount per serving")
    yp -= 7
    
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x + PADDING, yp, "Calories")
    c.drawString(x + PADDING + 45, yp, str(p.get('Calories', '')))
    yp -= 2
    
    c.setLineWidth(2.5)
    c.line(x + PADDING, yp, x + cfg['width'] - PADDING, yp)
    yp -= 5
    
    c.setFont("Helvetica-Bold", 5)
    c.drawRightString(x + cfg['width'] - PADDING, yp, "% Daily Value*")
    yp -= 7
    
    c.setLineWidth(0.3)
    
    def dn(label, campo, bold=True, indent=0):
        nonlocal yp
        valor = p.get(campo, '')
        dv = calc_dv(campo, valor)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 5)
        c.drawString(x + PADDING + indent, yp, f"{label} {valor}")
        if dv:
            c.drawRightString(x + cfg['width'] - PADDING, yp, f"{dv}%")
        yp -= 1
        c.line(x + PADDING, yp, x + cfg['width'] - PADDING, yp)
        yp -= 6
    
    dn("Total Fat", "Total_Fat")
    dn("Saturated Fat", "Saturated_Fat", False, 7)
    dn("Trans Fat", "Trans_Fat", False, 7)
    dn("Cholesterol", "Cholesterol")
    dn("Sodium", "Sodium")
    dn("Total Carb.", "Total_Carbohydrate")
    dn("Dietary Fiber", "Dietary_Fiber", False, 7)
    dn("Total Sugars", "Total_Sugars", False, 7)
    dn("  Incl. Added Sugars", "Added_Sugars", False, 10)
    
    c.setFont("Helvetica-Bold", 5)
    pv = p.get('Protein', '')
    pd = calc_dv('Protein', pv)
    c.drawString(x + PADDING, yp, f"Protein {pv}")
    if pd:
        c.drawRightString(x + cfg['width'] - PADDING, yp, f"{pd}%")
    yp -= 1
    
    c.setLineWidth(2.5)
    c.line(x + PADDING, yp, x + cfg['width'] - PADDING, yp)
    c.setLineWidth(0.3)
    yp -= 5
    
    c.setFont("Helvetica", 5)
    for campo, label in [('Vitamin_D', 'Vitamin D'), ('Calcium', 'Calcium'), 
                          ('Iron', 'Iron'), ('Potassium', 'Potassium')]:
        val = p.get(campo, '')
        if val and str(val).strip() != 'None':
            dv = calc_dv(campo, val)
            c.drawString(x + PADDING, yp, f"{label} {val}")
            if dv:
                c.drawRightString(x + cfg['width'] - PADDING, yp, f"{dv}%")
            yp -= 6
    
    c.setFont("Helvetica", 3.5)
    texto = "* The % Daily Value tells you how much a nutrient in a serving of food contributes to a daily diet. 2000 calories a day is used for general nutrition advice."
    max_width = cfg['width'] - 2 * PADDING
    linea = ""
    for palabra in texto.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 3.5) < max_width:
            linea = test
        else:
            c.drawString(x + PADDING, yp, linea.strip())
            yp -= 4.5
            linea = palabra + " "
    if linea:
        c.drawString(x + PADDING, yp, linea.strip())
        yp -= 5
    
    c.setFont("Helvetica-Bold", 5)
    c.drawString(x + PADDING, yp, "INGREDIENTS:")
    yp -= 5
    
    c.setFont("Helvetica", 4.5)
    ingredientes = str(p.get('Ingredients', ''))
    linea = ""
    for palabra in ingredientes.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 4.5) < max_width:
            linea = test
        else:
            c.drawString(x + PADDING, yp, linea.strip())
            yp -= 5
            linea = palabra + " "
    if linea:
        c.drawString(x + PADDING, yp, linea.strip())
        yp -= 6
    
    c.setFont("Helvetica-Bold", 5)
    allergens = p.get('Allergens', '')
    if allergens and str(allergens) != "NONE":
        c.drawString(x + PADDING, yp, allergens)
        yp -= 6
    
    c.setFont("Helvetica-Bold", 5.5)
    c.drawString(x + PADDING, yp, f"EXP: {p.get('Expiration_Date', '')}")
    yp -= 8
    
    c.setFont("Helvetica", 4.5)
    
    imported_by = p.get('Imported_By', '')
    has_imported = imported_by and str(imported_by).strip() and str(imported_by).strip() != 'None'
    
    if has_imported:
        ancho_total = cfg['width'] - 2 * PADDING
        ancho_cuadrito = ancho_total / 2 - 5
        
        left_x = x + PADDING
        c.drawString(left_x, yp, "Distributed by:")
        c.drawString(left_x, yp - 5, "Amaya Express")
        c.drawString(left_x, yp - 10, "472 Somerset St.,")
        c.drawString(left_x, yp - 15, "N. Plainfield, NJ 07060")
        
        right_x = x + PADDING + ancho_cuadrito + 10
        imported_parts = str(imported_by).split(',')
        
        c.drawString(right_x, yp, "Imported by:")
        if len(imported_parts) >= 1:
            c.drawString(right_x, yp - 5, imported_parts[0].strip())
        if len(imported_parts) >= 2:
            c.drawString(right_x, yp - 10, imported_parts[1].strip())
        if len(imported_parts) >= 3:
            c.drawString(right_x, yp - 15, imported_parts[2].strip())
    else:
        c.drawString(x + PADDING, yp, "Distributed by: Amaya Express")
        c.drawString(x + PADDING, yp - 5, "472 Somerset St., N. Plainfield, NJ 07060")

def dibujar_pls(c, x, y, p, cfg):
    """PLS504"""
    yp = y + cfg['height'] - PADDING * 0.5
    c.setFillColor(AMAYA_BLUE)
    c.setStrokeColor(black)
    
    ls = 0.3 * inch
    if os.path.exists(LOGO_PATH):
        c.drawImage(LOGO_PATH, x + PADDING * 0.5, yp - ls, ls, ls, preserveAspectRatio=True)
    yp -= ls + 2
    
    c.setFont("Helvetica-Bold", 7)
    nombre = p.get('Product_Name', '')
    tw = c.stringWidth(nombre, "Helvetica-Bold", 7)
    c.drawString(x + (cfg['width'] - tw) / 2, yp, nombre)
    yp -= 7
    
    nombre_eng = p.get('Product_Name_English', '')
    if nombre_eng and str(nombre_eng).strip() != 'None':
        c.setFont("Helvetica", 6)
        tw = c.stringWidth(nombre_eng, "Helvetica", 6)
        c.drawString(x + (cfg['width'] - tw) / 2, yp, nombre_eng)
    yp -= 8
    
    c.setFont("Helvetica", 5)
    c.drawString(x + PADDING * 0.5, yp, f"Wt: {p.get('Net_Weight', '')}")
    c.drawRightString(x + cfg['width'] - PADDING * 0.5, yp, f"EXP: {p.get('Expiration_Date', '')}")
    yp -= 7
    
    c.setFont("Helvetica", 4)
    c.drawString(x + PADDING * 0.5, yp, "Amaya Express")
    
    imported_by = p.get('Imported_By', '')
    if imported_by and str(imported_by).strip() != 'None':
        tw = c.stringWidth(imported_by, "Helvetica", 4)
        c.drawString(x + cfg['width'] - PADDING * 0.5 - tw, yp, imported_by)

def wrap_texto(c, texto, fuente, tam, ancho_max, xi, ypos, interlinea, dibujar=True):
    """
    Dibuja texto con salto de línea automático. Devuelve la posición Y final.
    Con dibujar=False no escribe nada (solo calcula el alto que ocuparía) - se usa
    para medir el bloque de texto antes de centrarlo verticalmente.
    """
    c.setFont(fuente, tam)
    linea = ""
    for palabra in str(texto).split():
        test = linea + palabra + " "
        if c.stringWidth(test, fuente, tam) < ancho_max:
            linea = test
        else:
            if dibujar:
                c.drawString(xi, ypos, linea.strip())
            ypos -= interlinea
            linea = palabra + " "
    if linea:
        if dibujar:
            c.drawString(xi, ypos, linea.strip())
        ypos -= interlinea
    return ypos

def dibujar_avery5260(c, x, y, p, cfg):
    """
    AVERY 5260 - 1" x 2-5/8" - Etiqueta de identificación + Nutrition Facts compacta.
    Esta etiqueta NO tiene que cumplir la restricción de tamaño mínimo de fuente de la
    FDA (21 CFR 101.9) porque no se usa para ese fin - por eso el texto puede ir más
    chico que en AVERY_8164/LACTEO_AVERY_8164.

    Layout: columna izquierda (nombre, ingredientes, distribuidor/importador, peso/EXP,
    contacto) + columna derecha con Nutrition Facts a todo el alto de la etiqueta.

    Los nutrientes en 0 (o vacíos) se omiten (ver es_cero()) y las filas que sí quedan
    se reparten en partes iguales para ocupar todo el alto disponible de la etiqueta,
    tal como se pidió.
    """
    c.setFillColor(black)
    c.setStrokeColor(black)

    # AJUSTE (feedback de impresión de prueba): se acercan los dos grupos para
    # reducir el espacio vacío del medio, sin tocar márgenes de la hoja (margin_left/
    # h_spacing en LABEL_CONFIGS). El texto de la izquierda se corre ~1.5mm a la
    # derecha (0.04" -> 0.10") y la columna de Nutrition Facts se ensancha un poco
    # hacia la izquierda (0.82" -> 0.90"), sin moverla del borde derecho de la etiqueta.
    # AJUSTE 2 (línea derecha): se agrega una línea vertical simétrica a la derecha
    # de Nutrition Facts. Para que no quede pegada al borde de la etiqueta (y no se
    # corte al recortar/imprimir), todo el bloque de Nutrition Facts se recorre un
    # poco a la izquierda dejando un margen derecho (right_margin) para esa línea.
    nf_width = 0.90 * inch
    gap = 0.05 * inch
    right_margin = 0.07 * inch
    nf_x = x + cfg['width'] - right_margin - nf_width
    left_x = x + 0.10 * inch
    left_width = cfg['width'] - nf_width - gap - 0.08 * inch

    # ---------- COLUMNA IZQUIERDA ----------
    # AJUSTE (feedback de impresión real): "Coco Rallado" quedaba pegado al borde
    # superior y se cortaba un poco al despegar la etiqueta, mientras sobraba espacio
    # abajo. Ahora el bloque se mide primero (sin dibujar) y luego se dibuja centrado
    # verticalmente dentro de la etiqueta, dejando el mismo margen arriba y abajo.
    def dibujar_columna_izquierda(yp_inicial, dibujar=True):
        yp = yp_inicial

        nombre = str(p.get('Product_Name', '') or '').strip()
        if dibujar:
            c.setFont("Helvetica-Bold", 6.5)
            c.drawString(left_x, yp, nombre)
        yp -= 6.5

        nombre_eng = p.get('Product_Name_English', '')
        if nombre_eng and str(nombre_eng).strip() not in ('', 'None'):
            if dibujar:
                c.setFont("Helvetica-Oblique", 5)
                c.drawString(left_x, yp, str(nombre_eng).strip())
            yp -= 6

        ingredientes = p.get('Ingredients', '')
        if ingredientes and str(ingredientes).strip() not in ('', 'None'):
            if dibujar:
                c.setFont("Helvetica-Bold", 4.2)
                c.drawString(left_x, yp, "INGREDIENTS:")
            yp -= 4.5
            yp = wrap_texto(c, ingredientes, "Helvetica", 4, left_width, left_x, yp, 4.2, dibujar)

        yp -= 1
        if dibujar:
            c.setFont("Helvetica-Bold", 4)
            c.drawString(left_x, yp, "Distributed by:")
        yp -= 4
        if dibujar:
            c.setFont("Helvetica", 3.7)
            c.drawString(left_x, yp, "Amaya Express Int'l, 472 Somerset St,")
        yp -= 3.9
        if dibujar:
            c.drawString(left_x, yp, "North Plainfield, NJ 06070")
        yp -= 4.3

        imported_by = p.get('Imported_By', '')
        if imported_by and str(imported_by).strip() not in ('', 'None'):
            if dibujar:
                c.setFont("Helvetica-Bold", 4)
                c.drawString(left_x, yp, "Imported by:")
            yp -= 4
            if dibujar:
                c.setFont("Helvetica", 3.7)
            for parte in str(imported_by).split(',')[:2]:
                if dibujar:
                    c.drawString(left_x, yp, parte.strip())
                yp -= 3.9

        net_weight = p.get('Net_Weight', '')
        exp = p.get('Expiration_Date', '')
        pie = []
        if net_weight and str(net_weight).strip() not in ('', 'None'):
            pie.append(f"Net Wt: {net_weight}")
        if exp and str(exp).strip() not in ('', 'None'):
            pie.append(f"EXP: {formatear_exp(exp)}")
        if pie:
            yp -= 1
            if dibujar:
                c.setFont("Helvetica-Bold", 4.2)
                c.drawString(left_x, yp, "   ".join(pie))
            yp -= 4.5

        if dibujar:
            c.setFont("Helvetica", 3.6)
            c.drawString(left_x, yp, "(908) 405-5553 / (908) 405-3072")
        yp -= 3.8
        if dibujar:
            c.drawString(left_x, yp, "amayaexpress21@gmail.com")

        return yp

    # 1) Medir el alto real del bloque (sin dibujar nada)
    yp_final_medido = dibujar_columna_izquierda(0, dibujar=False)
    alto_columna_izquierda = 0 - yp_final_medido

    # 2) Margen simétrico arriba/abajo dentro de la etiqueta (mínimo 2pt de seguridad)
    margen_vertical = max((cfg['height'] - alto_columna_izquierda) / 2, 2)

    # 3) Dibujar de verdad, ya centrado
    dibujar_columna_izquierda(y + cfg['height'] - margen_vertical, dibujar=True)

    # ---------- COLUMNA NUTRITION FACTS (todo el alto, como estaba antes) ----------
    # NOTA: se probó centrarla con el mismo margen que la columna izquierda, pero no
    # se vio bien (obligaba a achicar más la letra) - por pedido del usuario, esta
    # columna vuelve a usar todo el alto de la etiqueta (borde a borde), solo la
    # columna izquierda quedó centrada.
    nf_top = y + cfg['height']
    nf_bottom = y
    linea_vertical_top = nf_top - 1.5  # se define aquí arriba; el final se calcula
                                        # después de dibujar las filas (línea_vertical_bottom)

    # Línea horizontal de arriba, cerrando el cuadro junto con las dos verticales
    # (izquierda y derecha) que ya se dibujan más abajo, a la misma altura donde
    # empiezan esas verticales.
    c.setLineWidth(0.4)
    c.line(nf_x - gap / 2, linea_vertical_top,
           nf_x + nf_width + right_margin / 2, linea_vertical_top)

    yp_nf = nf_top - 6.5
    c.setFont("Helvetica-Bold", 6)
    c.drawString(nf_x, yp_nf, "Nutrition Facts")
    yp_nf -= 6

    serving_size = p.get('Serving_Size', '')
    if serving_size and str(serving_size).strip() not in ('', 'None'):
        c.setFont("Helvetica", 3.6)
        c.drawString(nf_x, yp_nf, f"Serving Size {serving_size}")
        yp_nf -= 4

    c.setLineWidth(1)
    c.line(nf_x, yp_nf, nf_x + nf_width, yp_nf)
    yp_nf -= 4.5

    calories = p.get('Calories', '')
    if calories and str(calories).strip() not in ('', 'None'):
        c.setFont("Helvetica-Bold", 6)
        c.drawString(nf_x, yp_nf, "Calories")
        c.drawString(nf_x + 32, yp_nf, str(calories))
        yp_nf -= 2
        c.setLineWidth(1.8)
        c.line(nf_x, yp_nf, nf_x + nf_width, yp_nf)
        yp_nf -= 4

    c.setFont("Helvetica-Bold", 3.3)
    c.drawRightString(nf_x + nf_width, yp_nf, "% DV*")
    yp_nf -= 3.5

    header_bottom = yp_nf  # a partir de aquí se reparten las filas de nutrientes

    filas = [
        ("Total Fat", "Total_Fat", True, 0),
        ("Sat. Fat", "Saturated_Fat", False, 4),
        ("Trans Fat", "Trans_Fat", False, 4),
        ("Cholesterol", "Cholesterol", True, 0),
        ("Sodium", "Sodium", True, 0),
        ("Total Carb.", "Total_Carbohydrate", True, 0),
        ("Fiber", "Dietary_Fiber", False, 4),
        ("Sugars", "Total_Sugars", False, 4),
        ("  Added Sugars", "Added_Sugars", False, 6),
        ("Protein", "Protein", True, 0),
        ("Vitamin D", "Vitamin_D", False, 0),
        ("Calcium", "Calcium", False, 0),
        ("Iron", "Iron", False, 0),
        ("Potassium", "Potassium", False, 0),
    ]
    filas_visibles = [f for f in filas if not es_cero(p.get(f[1], ''))]

    n = len(filas_visibles)
    if n:
        espacio_disponible = max(header_bottom - nf_bottom - 2, 0)

        # Cada fila necesita un mínimo de espacio para que el texto no choque con
        # la línea divisoria de abajo. Empezamos en 3.6pt (tamaño "normal" para
        # esta etiqueta) y solo achicamos la letra si con muchos nutrientes no
        # alcanza el alto disponible; nunca bajamos de 2.6pt (deja de ser legible).
        tam_fuente = 3.6
        pre_linea = 0.9      # espacio fijo entre el texto y su línea divisoria
        post_linea_min = tam_fuente  # espacio mínimo entre una línea y el texto siguiente
        while n * (pre_linea + post_linea_min) > espacio_disponible and tam_fuente > 2.6:
            tam_fuente -= 0.1
            post_linea_min = tam_fuente

        paso = espacio_disponible / n
        post_linea = max(post_linea_min, paso - pre_linea)

        y_fila = header_bottom
        for label, campo, bold, indent in filas_visibles:
            valor = p.get(campo, '')
            dv = calc_dv(campo, valor)
            c.setFont("Helvetica-Bold" if bold else "Helvetica", tam_fuente)
            c.drawString(nf_x + indent, y_fila, f"{label} {valor}")
            if dv:
                c.drawRightString(nf_x + nf_width, y_fila, f"{dv}%")
            y_fila -= pre_linea
            c.setLineWidth(0.3)
            c.line(nf_x, y_fila, nf_x + nf_width, y_fila)
            y_fila -= post_linea
        linea_vertical_bottom = y_fila + post_linea  # justo debajo de la última línea dibujada
    else:
        linea_vertical_bottom = header_bottom

    # Separadores verticales (izquierda y derecha de Nutrition Facts): SOLO del alto
    # del contenido real de esta etiqueta (título hasta la última fila), no del alto
    # completo de la etiqueta. Antes llegaban de borde a borde y, como las etiquetas
    # se tocan verticalmente (v_spacing 0), las líneas de etiquetas consecutivas se
    # unían formando una sola raya en toda la hoja.
    # Más delgadas que antes (0.75 -> 0.4) según feedback de la prueba impresa.
    c.setLineWidth(0.4)
    c.line(nf_x - gap / 2, linea_vertical_bottom, nf_x - gap / 2, linea_vertical_top)
    c.line(nf_x + nf_width + right_margin / 2, linea_vertical_bottom,
           nf_x + nf_width + right_margin / 2, linea_vertical_top)

def generar_pdf(rex, rpdf):
    print(f"\n📖 Leyendo {rex}...")
    try:
        wb = openpyxl.load_workbook(rex)
        s = wb["Productos_FDA"]
    except Exception as e:
        print(f"❌ {e}")
        return False
    
    h = [s.cell(1, c).value for c in range(1, 50) if s.cell(1, c).value]
    ps = []
    r = 2
    while True:
        n = s.cell(r, 1).value
        if not n or not str(n).strip():
            break
        ps.append({h[c-1]: s.cell(r, c).value or "" for c in range(1, len(h)+1)})
        r += 1
    
    wb.close()
    
    lt = ps[0].get('Label_Type', 'AVERY_8164') if ps else 'AVERY_8164'
    cfg = LABEL_CONFIGS.get(lt, LABEL_CONFIGS['AVERY_8164'])
    tot, hjs = len(ps), (len(ps) + cfg['per_page'] - 1) // cfg['per_page']
    
    print(f"✅ {tot} etiquetas - {hjs} hojas")
    print(f"🎯 {lt}\n🎨 Generando...\n")
    
    # Configuración para máxima compatibilidad con impresoras
    c = canvas.Canvas(rpdf, pagesize=letter, 
                     invariant=0,        # CAMBIADO: cada PDF único (evita caché)
                     pageCompression=0)  # Sin compresión
    
    # Agregar metadata única para forzar que cada PDF sea diferente
    import datetime
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S_%f")
    c.setAuthor(f"Amaya_{timestamp}")
    c.setTitle(f"Etiquetas_FDA_{timestamp}")
    
    pos = []
    for f in range(cfg['rows']):
        for cl in range(cfg['columns']):
            pos.append((cfg['margin_left'] + cl * (cfg['width'] + cfg['h_spacing']),
                       PAGE_HEIGHT - cfg['margin_top'] - (f + 1) * cfg['height'] - f * cfg['v_spacing']))
    
    for i, p in enumerate(ps):
        if i % cfg['per_page'] == 0 and i > 0:
            c.showPage()
        x, y = pos[i % cfg['per_page']]
        
        if lt == 'LACTEO_AVERY_8164':
            dibujar_lacteo_avery(c, x, y, p, cfg)
        elif lt == 'PLS504':
            dibujar_pls(c, x, y, p, cfg)
        elif lt == 'AVERY_5260':
            dibujar_avery5260(c, x, y, p, cfg)
        else:
            dibujar_avery(c, x, y, p, cfg)
        
        if (i + 1) % cfg['per_page'] == 0 or i + 1 == tot:
            print(f"   {i + 1}/{tot}")
    
    c.save()
    print(f"\n✅ PDF: {rpdf}\n   {tot} etiquetas en {hjs} hojas")
    return True

if __name__ == "__main__":
    print("=" * 70 + "\n🏷️  AMAYA EXPRESS - VERSIÓN CORREGIDA\n" + "=" * 70)
    ex, pdf = "Etiquetas_Para_Imprimir.xlsx", "Etiquetas_Para_Imprimir.pdf"
    if not os.path.exists(ex):
        print(f"\n❌ No existe {ex}")
        sys.exit(1)
    if generar_pdf(ex, pdf):
        print("\n" + "=" * 70 + "\n✅ LISTO\n" + "=" * 70)
    else:
        print("\n❌ Error")
        sys.exit(1)
