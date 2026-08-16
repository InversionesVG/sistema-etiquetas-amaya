#!/usr/bin/env python3
"""
Sistema de Generación de Etiquetas FDA
--------------------------------------
Este script:
1. Lee el Excel original con todos los productos
2. Levanta un servidor web local
3. Sirve la interfaz HTML para seleccionar productos
4. Genera Excel temporal con productos duplicados según cantidad
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import json
import sys
from http.server import HTTPServer, BaseHTTPRequestHandler
from urllib.parse import parse_qs
import os

# ============================================================================
# CONFIGURACIÓN
# ============================================================================

EXCEL_ORIGINAL = "Base_Datos_Etiquetas_FDA.xlsx"
EXCEL_TEMPORAL = "Etiquetas_Para_Imprimir.xlsx"
HTML_FILE = "Selector_Etiquetas.html"
PORT = 8000

# ============================================================================
# FUNCIONES PARA LEER EXCEL
# ============================================================================

def leer_productos_desde_excel(ruta_excel):
    """
    Lee todos los productos del Excel original.
    
    Analogía con VB6:
    -----------------
    En VB6 harías:
        Set rs = db.Execute("SELECT * FROM Productos")
        While Not rs.EOF
            ' procesar registro
            rs.MoveNext
        Wend
    
    Aquí es lo mismo: abrimos el Excel, recorremos filas,
    guardamos cada producto en una lista.
    """
    print(f"📖 Leyendo productos desde {ruta_excel}...")
    
    try:
        wb = openpyxl.load_workbook(ruta_excel)
        sheet = wb["Productos_FDA"]
    except FileNotFoundError:
        print(f"❌ ERROR: No se encontró el archivo {ruta_excel}")
        print(f"   Asegúrate de que esté en la misma carpeta que este script")
        sys.exit(1)
    except KeyError:
        print(f"❌ ERROR: El Excel no tiene la hoja 'Productos_FDA'")
        sys.exit(1)
    
    # Leer encabezados (fila 1) - lee TODAS las columnas que existan
    headers = []
    for col in range(1, 50):  # Lee hasta 50 columnas (más que suficiente)
        cell_value = sheet.cell(row=1, column=col).value
        if cell_value:
            headers.append(cell_value)
        elif col > 10 and not cell_value:  # Si después de la col 10 encuentra vacío, termina
            break
    
    # Leer productos (desde fila 2 en adelante)
    productos = []
    row = 2
    
    while True:
        # Leer nombre del producto (columna A)
        nombre = sheet.cell(row=row, column=1).value
        
        # Si no hay nombre, terminamos (llegamos al final)
        if not nombre or nombre.strip() == "":
            break
        
        # Crear diccionario con todos los datos del producto
        producto = {}
        for col_index, header in enumerate(headers, 1):
            valor = sheet.cell(row=row, column=col_index).value
            producto[header] = str(valor) if valor is not None else ""
        
        productos.append(producto)
        row += 1
    
    wb.close()
    
    print(f"✅ Se encontraron {len(productos)} productos")
    return productos


def generar_excel_temporal(productos_seleccionados, ruta_salida, label_type='AVERY_8164'):
    """
    Genera Excel temporal duplicando productos según cantidad.
    
    IMPORTANTE: Ahora copia TODAS las columnas del Excel original,
    incluyendo Product_Name_English e Imported_By.
    También guarda el tipo de etiqueta seleccionado.
    """
    print(f"\n📝 Generando Excel temporal...")
    print(f"   Tipo de etiqueta: {label_type}")
    
    wb = openpyxl.Workbook()
    sheet = wb.active
    sheet.title = "Productos_FDA"
    
    # Estilos
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # IMPORTANTE: Obtener headers del primer producto (que tiene TODAS las columnas)
    if not productos_seleccionados:
        print("❌ No hay productos seleccionados")
        return None
    
    headers = list(productos_seleccionados[0].keys())
    
    # Agregar columna para tipo de etiqueta si no existe
    if 'Label_Type' not in headers:
        headers.append('Label_Type')
    
    # Crear encabezados
    for col_num, header in enumerate(headers, 1):
        cell = sheet.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = border
    
    # Ajustar anchos de columna (ancho por defecto)
    for i in range(1, len(headers) + 1):
        col_letter = openpyxl.utils.get_column_letter(i)
        sheet.column_dimensions[col_letter].width = 15
    
    # Anchos especiales para columnas conocidas
    special_widths = {
        'Product_Name': 30,
        'Product_Name_English': 30,
        'Ingredients': 60,
        'Allergens': 35,
        'Imported_By': 40
    }
    
    for col_num, header in enumerate(headers, 1):
        if header in special_widths:
            col_letter = openpyxl.utils.get_column_letter(col_num)
            sheet.column_dimensions[col_letter].width = special_widths[header]
    
    # Duplicar productos según cantidad
    current_row = 2
    total_etiquetas = 0
    
    for producto in productos_seleccionados:
        cantidad = int(producto.get("Cantidad_a_Imprimir", 1))
        total_etiquetas += cantidad
        
        print(f"   • {producto.get('Product_Name', 'Sin nombre')}: {cantidad} etiquetas")
        
        # Duplicar el producto 'cantidad' veces
        for _ in range(cantidad):
            for col_index, header in enumerate(headers, 1):
                cell = sheet.cell(row=current_row, column=col_index)
                # Si es la columna Label_Type, poner el tipo seleccionado
                if header == 'Label_Type':
                    cell.value = label_type
                else:
                    cell.value = producto.get(header, "")
                cell.border = border
                cell.alignment = Alignment(vertical="center")
            current_row += 1
    
    # Guardar
    wb.save(ruta_salida)
    wb.close()
    
    # Calcular hojas según tipo de etiqueta
    if label_type == 'PLS504':
        hojas = (total_etiquetas + 9) // 10  # 10 etiquetas por hoja
    elif label_type == 'AVERY_5260':
        hojas = (total_etiquetas + 29) // 30  # 30 etiquetas por hoja (Dulces)
    else:  # AVERY_8164 / LACTEO_AVERY_8164
        hojas = (total_etiquetas + 5) // 6   # 6 etiquetas por hoja
    
    print(f"\n✅ Excel temporal generado exitosamente:")
    print(f"   📄 Archivo: {ruta_salida}")
    print(f"   🏷️  Total de etiquetas: {total_etiquetas}")
    print(f"   📋 Hojas a imprimir: {hojas}")
    print(f"   📊 Columnas incluidas: {len(headers)}")
    print(f"   🎯 Tipo: {label_type}")
    
    return ruta_salida
    print(f"   📋 Hojas a imprimir: {hojas}")
    
    return ruta_salida


# ============================================================================
# SERVIDOR WEB
# ============================================================================

class EtiquetasHandler(BaseHTTPRequestHandler):
    """
    Servidor web simple que maneja las peticiones del navegador.
    
    Analogía con VB6:
    -----------------
    En VB6, cuando hacías un formulario, el usuario hacía clic
    en botones y el código respondía (eventos Click).
    
    Aquí es similar:
    - El navegador solicita archivos (GET)
    - El navegador envía datos (POST)
    - Este servidor responde
    """
    
    def do_GET(self):
        """Manejar peticiones GET (cargar archivos)"""
        
        if self.path == '/':
            # Servir HTML principal
            self.send_response(200)
            self.send_header('Content-type', 'text/html; charset=utf-8')
            self.end_headers()
            
            with open(HTML_FILE, 'rb') as f:
                self.wfile.write(f.read())
        
        elif self.path == '/productos.json':
            # Servir lista de productos en JSON
            self.send_response(200)
            self.send_header('Content-type', 'application/json; charset=utf-8')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            productos = leer_productos_desde_excel(EXCEL_ORIGINAL)
            self.wfile.write(json.dumps(productos, ensure_ascii=False).encode('utf-8'))
        
        else:
            self.send_response(404)
            self.end_headers()
    
    def do_POST(self):
        """Manejar peticiones POST (generar Excel)"""
        
        if self.path == '/generate':
            # Leer datos enviados desde el navegador
            content_length = int(self.headers['Content-Length'])
            post_data = self.rfile.read(content_length)
            data = json.loads(post_data.decode('utf-8'))
            
            productos_seleccionados = data.get('products', [])
            label_type = data.get('label_type', 'AVERY_8164')  # Default: Avery 8164
            
            # Generar Excel temporal
            ruta_excel = generar_excel_temporal(productos_seleccionados, EXCEL_TEMPORAL, label_type)
            
            # Enviar archivo Excel al navegador para descarga
            self.send_response(200)
            self.send_header('Content-type', 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            self.send_header('Content-Disposition', f'attachment; filename="{EXCEL_TEMPORAL}"')
            self.send_header('Access-Control-Allow-Origin', '*')
            self.end_headers()
            
            with open(ruta_excel, 'rb') as f:
                self.wfile.write(f.read())
    
    def do_OPTIONS(self):
        """Manejar preflight CORS"""
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')
        self.end_headers()
    
    def log_message(self, format, *args):
        """Silenciar logs del servidor (opcional)"""
        pass


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================

def main():
    """Función principal del sistema"""
    
    print("=" * 70)
    print("🏷️  SISTEMA DE GENERACIÓN DE ETIQUETAS FDA")
    print("=" * 70)
    print()
    
    # Verificar que existan los archivos necesarios
    if not os.path.exists(EXCEL_ORIGINAL):
        print(f"❌ ERROR: No se encontró {EXCEL_ORIGINAL}")
        print(f"   Coloca este script en la misma carpeta que tu Excel")
        sys.exit(1)
    
    if not os.path.exists(HTML_FILE):
        print(f"❌ ERROR: No se encontró {HTML_FILE}")
        print(f"   Descarga todos los archivos del sistema")
        sys.exit(1)
    
    # Cargar productos para verificar
    productos = leer_productos_desde_excel(EXCEL_ORIGINAL)
    
    if len(productos) == 0:
        print("⚠️  ADVERTENCIA: El Excel no tiene productos")
        print("   Agrega productos al Excel antes de continuar")
        sys.exit(1)
    
    print()
    print("🌐 Iniciando servidor web...")
    print(f"   Puerto: {PORT}")
    print()
    print("=" * 70)
    print("📱 INSTRUCCIONES:")
    print("=" * 70)
    print()
    print(f"1. Abre tu navegador (Chrome, Safari, Firefox)")
    print(f"2. Ve a la dirección: http://localhost:{PORT}")
    print(f"3. Selecciona los productos que deseas imprimir")
    print(f"4. Define la cantidad de etiquetas de cada uno")
    print(f"5. Clic en 'Generar Excel para Impresión'")
    print(f"6. Se descargará '{EXCEL_TEMPORAL}'")
    print(f"7. Vincula ESE Excel con Word y listo!")
    print()
    print("Para DETENER el servidor: presiona Ctrl+C")
    print("=" * 70)
    print()
    
    # Iniciar servidor
    try:
        server = HTTPServer(('localhost', PORT), EtiquetasHandler)
        print(f"✅ Servidor corriendo en http://localhost:{PORT}")
        print()
        server.serve_forever()
    except KeyboardInterrupt:
        print("\n\n🛑 Servidor detenido por el usuario")
        server.shutdown()
    except OSError as e:
        if "Address already in use" in str(e):
            print(f"\n❌ ERROR: El puerto {PORT} ya está en uso")
            print(f"   Cierra otros programas que usen ese puerto o cambia PORT en el script")
        else:
            print(f"\n❌ ERROR: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
