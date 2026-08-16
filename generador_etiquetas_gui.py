#!/usr/bin/env python3
"""
GENERADOR DE ETIQUETAS AMAYA EXPRESS
Aplicación con interfaz gráfica para Windows
Versión: 1.0
"""

import sys
import os
import platform
from datetime import datetime
from PyQt5.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QLabel, QPushButton, QCheckBox, QSpinBox, QComboBox,
    QScrollArea, QFrame, QMessageBox, QGroupBox, QGridLayout, QLineEdit
)
from PyQt5.QtCore import Qt
from PyQt5.QtGui import QFont, QIcon, QPixmap

# Importar las librerías para PDF
import openpyxl
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.pdfgen import canvas
from reportlab.lib.colors import HexColor, black
import re

# ============================================================================
# INFORMACIÓN DE VERSIÓN
# ============================================================================
VERSION = "1.3.0"  # Formato: MAYOR.MENOR.PARCHE
VERSION_DATE = "Agosto 2026"

# ============================================================================
# CONFIGURACIÓN DE RUTAS (MULTIPLATAFORMA)
# ============================================================================
# Detectar sistema operativo y configurar rutas apropiadas
if platform.system() == "Windows":
    BASE_DIR = "C:/Sistema_Etiquetas"
else:  # Mac/Linux
    # AJUSTE v1.3.0: unificado con la carpeta única del proyecto (antes era
    # ~/Documents/Sistema_Etiquetas, separada del código fuente). En Windows
    # sigue igual (C:/Sistema_Etiquetas) - no se tocó esa parte a propósito.
    BASE_DIR = os.path.expanduser("~/Documents/Python/Amaya_Labels")

# Crear directorio si no existe
if not os.path.exists(BASE_DIR):
    os.makedirs(BASE_DIR)

EXCEL_PATH = os.path.join(BASE_DIR, "Base_Datos_Etiquetas_FDA.xlsx")
LOGO_PATH = os.path.join(BASE_DIR, "Amaya_Imagen.png")
FLAG_PATH = os.path.join(BASE_DIR, "Flag_El_Salvador.png")
FONDO_SUPERIOR_PATH = os.path.join(BASE_DIR, "Fondo_Superior.png")
FONDO_INFERIOR_PATH = os.path.join(BASE_DIR, "Fondo_Inferior.png")

# Rutas para San Julián
LOGO_SAN_JULIAN_PATH = os.path.join(BASE_DIR, "Logo_San_Julian_Vaca.png")
SELLO_SAN_JULIAN_PATH = os.path.join(BASE_DIR, "Sello_Verde_San_Julian.png")
FONDO_AZUL_SAN_JULIAN_PATH = os.path.join(BASE_DIR, "Fondo_Azul_San_Julian.png")

# ============================================================================
# CONFIGURACIÓN DE ETIQUETAS (de tu código original)
# ============================================================================
LABEL_CONFIGS = {
    'AVERY_8164': {
        'width': 3.33 * inch, 'height': 4 * inch, 'columns': 2, 'rows': 3, 'per_page': 6,
        'margin_left': 0.16 * inch, 'margin_top': 0.48 * inch, 
        'h_spacing': 0.85 * inch, 'v_spacing': -0.68 * inch,
    },
    'LACTEO_AVERY_8164': {
        'width': 4 * inch, 'height': 3.33 * inch, 'columns': 2, 'rows': 3, 'per_page': 6,
        'margin_left': 0.16 * inch, 'margin_top': 0.48 * inch,    
        'h_spacing': 0.18 * inch, 'v_spacing': 0.05 * inch,
    },
    'LACTEO_SAN_JULIAN': {
        'width': 4 * inch, 'height': 3.33 * inch, 'columns': 2, 'rows': 3, 'per_page': 6,
        'margin_left': 0.16 * inch, 'margin_top': 0.48 * inch,    
        'h_spacing': 0.18 * inch, 'v_spacing': 0.05 * inch,
    },
    'PLS504': {
        'width': 3 * inch, 'height': 2 * inch, 'columns': 2, 'rows': 5, 'per_page': 10,
        'margin_left': 1.20 * inch, 'margin_top': 0.3 * inch,
        'h_spacing': 0.15 * inch, 'v_spacing': 0 * inch,
    },
    'AVERY_5260': {
        # "Dulces" - Avery 5260: 1" x 2-5/8", 3 columnas x 10 filas = 30 por hoja
        # (specs oficiales Avery). Migrado desde el prototipo probado en Cowork
        # (agosto 2026) junto con Coco Rallado.
        'width': 2.625 * inch, 'height': 1 * inch, 'columns': 3, 'rows': 10, 'per_page': 30,
        'margin_left': 0.1875 * inch, 'margin_top': 0.5 * inch,
        'h_spacing': 0.125 * inch, 'v_spacing': 0 * inch,
    }
}

AMAYA_BLUE = HexColor('#0066B3')
DARK_BLUE = HexColor('#003366')
PAGE_WIDTH, PAGE_HEIGHT = letter
PADDING = 0.1 * inch

FDA_REF = {
    'Total_Fat': 78, 'Saturated_Fat': 20, 'Cholesterol': 300, 'Sodium': 2300,
    'Total_Carbohydrate': 275, 'Dietary_Fiber': 28, 'Added_Sugars': 50, 
    'Protein': 50, 'Vitamin_D': 20, 'Calcium': 1300, 'Iron': 18, 'Potassium': 4700
}

# ============================================================================
# FUNCIONES DE UTILIDAD (de tu código original)
# ============================================================================
def extraer_numero(t):
    m = re.search(r'(\d+\.?\d*)', str(t).strip()) if t and str(t).strip() else None
    return float(m.group(1)) if m else 0

def calc_dv(n, v):
    val = extraer_numero(v)
    return round((val / FDA_REF[n]) * 100) if n in FDA_REF and val > 0 else None

def es_cero(valor):
    """
    True si el valor está vacío/None, o si el primer número que contiene es 0.
    Se usa en AVERY_5260 (Dulces) para omitir filas de Nutrition Facts en 0
    (p.ej. "0g", "0mg") y así aprovechar el espacio para lo que sí tiene valor.
    NOTA: "<1g" NO se considera cero (el número detectado es 1), así que sí se muestra.
    """
    if valor is None:
        return True
    s = str(valor).strip()
    if not s or s == 'None':
        return True
    m = re.search(r'(\d+\.?\d*)', s)
    if not m:
        return True
    return float(m.group(1)) == 0

def formatear_exp(valor):
    """
    Formatea Expiration_Date como MM/AAAA (ej. "EXP:12/2027"). Si Excel guardó la
    celda como fecha, evita que salga con la hora pegada (ej. "2027-04-14 00:00:00").
    Si ya viene como texto, se deja tal cual.
    """
    if isinstance(valor, datetime):
        return valor.strftime('%m/%Y')
    return str(valor) if valor not in (None, '') else ''

# ============================================================================
# FUNCIONES DE DIBUJO (de tu código original - las incluiré después)
# ============================================================================
def dibujar_lacteo_avery(c, x, y, p, cfg):
    """Dibujar etiqueta láctea horizontal"""
    altura_superior = cfg['height'] * 0.6
    altura_inferior = cfg['height'] * 0.4
    
    # FONDOS
    if os.path.exists(FONDO_SUPERIOR_PATH):
        c.drawImage(FONDO_SUPERIOR_PATH, x, y + altura_inferior,
                   width=cfg['width'], height=altura_superior,
                   preserveAspectRatio=False, mask='auto')
    
    if os.path.exists(FONDO_INFERIOR_PATH):
        c.drawImage(FONDO_INFERIOR_PATH, x, y,
                   width=cfg['width'], height=altura_inferior,
                   preserveAspectRatio=False, mask='auto')
    
    # SECCIÓN SUPERIOR (60%)
    c.setFillColor(DARK_BLUE)
    yp = y + cfg['height'] - PADDING * 0.3
    yp -= 8
    c.setFont("Helvetica-Bold", 8)
    keep_text = "KEEP REFRIGERATED"
    text_width = c.stringWidth(keep_text, "Helvetica-Bold", 8)
    c.drawString(x + cfg['width'] - text_width - PADDING * 0.3, yp, keep_text)
    yp -= 15
    
    # Product_Name
    nombre_producto = p.get('Product_Name', '')
    c.setFont("Helvetica-Bold", 20)
    text_width = c.stringWidth(nombre_producto, "Helvetica-Bold", 20)
    c.drawString(x + (cfg['width'] - text_width) / 2, yp, nombre_producto)
    yp -= 16
    
    # Product_Name_English
    nombre_eng = p.get('Product_Name_English', '')
    if nombre_eng and str(nombre_eng).strip() != 'None':
        c.setFont("Helvetica-BoldOblique", 13)
        text_width = c.stringWidth(nombre_eng, "Helvetica-BoldOblique", 13)
        c.drawString(x + (cfg['width'] - text_width) / 2, yp, nombre_eng)
    
    # SECCIÓN INFERIOR (40%)
    y_inicio_azul = y + altura_inferior - PADDING * 0.3
    yp_azul = y_inicio_azul
    
    c.setFillColor(black)
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x + PADDING * 0.3, yp_azul, "INGREDIENTS:")
    yp_azul -= 6
    
    # Ingredientes
    c.setFont("Helvetica", 5)
    ingredientes = str(p.get('Ingredients', ''))
    max_width_ing = cfg['width'] - 1.3 * inch
    linea = ""
    
    for palabra in ingredientes.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 5) < max_width_ing:
            linea = test
        else:
            c.drawString(x + PADDING * 0.3, yp_azul, linea.strip())
            yp_azul -= 5
            linea = palabra + " "
    
    if linea:
        c.drawString(x + PADDING * 0.3, yp_azul, linea.strip())
        yp_azul -= 7
    
    # CONTAINS
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x + PADDING * 0.3, yp_azul, "CONTAINS: MILK PROTEINS.")
    yp_azul -= 8
    
    # FECHA DE VENCIMIENTO
    c.setFont("Helvetica-Bold", 6)
    exp_text = f"EXP: {p.get('Expiration_Date', '')}"
    c.drawString(x + PADDING * 0.3, yp_azul, exp_text)
    yp_azul -= 10
    
    # LOGO AMAYA
    if os.path.exists(LOGO_PATH):
        logo_size = 0.35 * inch
        c.drawImage(LOGO_PATH, x + PADDING * 0.3, yp_azul - logo_size,
                   logo_size, logo_size, preserveAspectRatio=True, mask='auto')
    
    # NET WEIGHT con fondo AZUL OSCURO
    c.setFillColor(DARK_BLUE)
    peso_x = x + PADDING * 0.3 + logo_size + 0.08 * inch
    peso_y = yp_azul - 0.35 * inch
    peso_ancho = 0.45 * inch
    peso_alto = 0.32 * inch
    
    c.rect(peso_x, peso_y, peso_ancho, peso_alto, fill=1, stroke=0)
    
    # Texto del peso en blanco
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 8)
    net_weight = str(p.get('Net_Weight', ''))
    
    if 'oz' in net_weight.lower():
        partes = net_weight.split('(')
        oz_text = partes[0].strip()
        gram_text = f"({partes[1]}" if len(partes) > 1 else ""
        
        tw = c.stringWidth(oz_text, "Helvetica-Bold", 8)
        c.drawString(peso_x + (peso_ancho - tw) / 2, peso_y + peso_alto - 11, oz_text)
        
        if gram_text:
            c.setFont("Helvetica-Bold", 6)
            tw = c.stringWidth(gram_text, "Helvetica-Bold", 6)
            c.drawString(peso_x + (peso_ancho - tw) / 2, peso_y + 4, gram_text)
    else:
        tw = c.stringWidth(net_weight, "Helvetica-Bold", 8)
        c.drawString(peso_x + (peso_ancho - tw) / 2, peso_y + peso_alto / 2 - 3, net_weight)
    
    # Distributed/Imported
    c.setFillColor(black)
    dist_x = peso_x + peso_ancho + 0.08 * inch
    dist_y = yp_azul - 0.05 * inch
    
    c.setFont("Helvetica", 4.5)
    c.drawString(dist_x, dist_y, "Distributed by:")
    c.drawString(dist_x, dist_y - 5, "Amaya Express International")
    c.drawString(dist_x, dist_y - 10, "472 Somerset St.,")
    c.drawString(dist_x, dist_y - 15, "North Plainfield, NJ 06070")
    
    imported_by = p.get('Imported_By', '')
    if imported_by and str(imported_by).strip() and str(imported_by).strip() != 'None':
        c.drawString(dist_x, dist_y - 22, "Imported by:")
        imported_parts = str(imported_by).split(',')
        y_offset = 27
        for part in imported_parts[:3]:
            c.drawString(dist_x, dist_y - y_offset, part.strip())
            y_offset += 5
    
    # Contacto
    contact_y = y + PADDING * 0.2
    c.setFont("Helvetica", 4.5)
    c.drawString(x + PADDING * 0.3, contact_y, "(908) 405-5553")
    
    email = "amayaexpress21@hotmail.com"
    email_width = c.stringWidth(email, "Helvetica", 4.5)
    c.drawString(x + cfg['width'] / 2 - email_width / 2, contact_y, email)
    
    phone2 = "(908) 405-3072"
    phone2_width = c.stringWidth(phone2, "Helvetica", 4.5)
    c.drawString(x + cfg['width'] - phone2_width - PADDING * 0.3, contact_y, phone2)
    
    # NUTRITION FACTS (compacto, solo hasta Protein)
    nf_x = x + cfg['width'] - 1.15 * inch
    nf_y = y + altura_inferior - PADDING * 0.5
    nf_width = 1.05 * inch
    
    c.setFillColor(black)
    c.setStrokeColor(black)
    
    # Borde
    c.setLineWidth(2)
    c.rect(nf_x, nf_y - 1.1 * inch, nf_width, 1.1 * inch, fill=0, stroke=1)
    c.setLineWidth(0.3)
    
    yp = nf_y - 8
    
    # Título
    c.setFont("Helvetica-Bold", 7)
    c.drawString(nf_x + PADDING * 0.3, yp, "Nutrition Facts")
    yp -= 1
    c.setLineWidth(0.5)
    c.line(nf_x + PADDING * 0.3, yp, nf_x + nf_width - PADDING * 0.3, yp)
    c.setLineWidth(0.3)
    yp -= 6
    
    # Servings
    c.setFont("Helvetica", 4.5)
    servings = p.get('Servings_Per_Container', '')
    if servings and str(servings).strip() != 'None':
        c.drawString(nf_x + PADDING * 0.3, yp, f"Servings per container: {servings}")
        yp -= 5
    
    # Serving Size
    serving_size = p.get('Serving_Size', '')
    c.setFont("Helvetica-Bold", 4.5)
    c.drawString(nf_x + PADDING * 0.3, yp, "Serving size")
    c.drawRightString(nf_x + nf_width - PADDING * 0.3, yp, str(serving_size))
    yp -= 1
    
    c.setLineWidth(2.5)
    c.line(nf_x + PADDING * 0.3, yp, nf_x + nf_width - PADDING * 0.3, yp)
    c.setLineWidth(0.3)
    yp -= 5
    
    # Calories
    c.setFont("Helvetica-Bold", 3.8)
    c.drawString(nf_x + PADDING * 0.3, yp, "Calories")
    c.setFont("Helvetica-Bold", 7)
    calories = p.get('Calories', '')
    c.drawRightString(nf_x + nf_width - PADDING * 0.3, yp, str(calories))
    yp -= 1
    
    c.setLineWidth(1.5)
    c.line(nf_x + PADDING * 0.3, yp, nf_x + nf_width - PADDING * 0.3, yp)
    c.setLineWidth(0.3)
    yp -= 4
    
    # % Daily Value
    c.setFont("Helvetica-Bold", 3.5)
    c.drawRightString(nf_x + nf_width - PADDING * 0.3, yp, "% Daily Value*")
    yp -= 1
    
    c.setLineWidth(1)
    c.line(nf_x + PADDING * 0.3, yp, nf_x + nf_width - PADDING * 0.3, yp)
    c.setLineWidth(0.3)
    yp -= 5
    
    # Función para dibujar nutriente
    def dn(label, field, bold=True, indent=0):
        nonlocal yp
        font = "Helvetica-Bold" if bold else "Helvetica"
        c.setFont(font, 3.8)
        valor = p.get(field, '')
        dv = calc_dv(field, valor)
        c.drawString(nf_x + PADDING * 0.3 + indent, yp, f"{label} {valor}")
        if dv:
            c.drawRightString(nf_x + nf_width - PADDING * 0.3, yp, f"{dv}%")
        yp -= 5
    
    # Nutrientes (solo hasta Protein)
    dn("Total Fat", "Total_Fat")
    dn("Saturated Fat", "Saturated_Fat", False, 7)
    dn("Trans Fat", "Trans_Fat", False, 7)
    dn("Cholesterol", "Cholesterol")
    dn("Sodium", "Sodium")
    dn("Total Carb.", "Total_Carbohydrate")
    dn("Dietary Fiber", "Dietary_Fiber", False, 7)
    dn("Total Sugars", "Total_Sugars", False, 7)
    dn("  Incl. Added Sugars", "Added_Sugars", False, 10)
    
    c.setFont("Helvetica-Bold", 5)
    pv = p.get('Protein', '')
    pd = calc_dv('Protein', pv)
    c.drawString(nf_x + PADDING * 0.3, yp, f"Protein {pv}")
    if pd:
        c.drawRightString(nf_x + nf_width - PADDING * 0.3, yp, f"{pd}%")


def dibujar_avery(c, x, y, p, cfg):
    """Versión normal Avery 8164 (vertical)"""
    yp = y + cfg['height'] - PADDING
    c.setFillColor(AMAYA_BLUE)
    c.setStrokeColor(black)
    
    logo_size = 0.4 * inch
    if os.path.exists(LOGO_PATH):
        c.drawImage(LOGO_PATH, x + PADDING, yp - logo_size, 
                   logo_size, logo_size, preserveAspectRatio=True)
    
    nombre_esp = p.get('Product_Name', '')
    nombre_eng = p.get('Product_Name_English', '')
    
    nombre_x_start = x + PADDING + logo_size + 0.05 * inch
    nombre_width = cfg['width'] - logo_size - PADDING - 0.05 * inch
    
    c.setFont("Helvetica-Bold", 9)
    tw = c.stringWidth(nombre_esp, "Helvetica-Bold", 9)
    c.drawString(nombre_x_start + (nombre_width - tw) / 2, yp - 0.12 * inch, nombre_esp)
    
    if nombre_eng and str(nombre_eng).strip() != 'None':
        c.setFont("Helvetica", 7)
        tw = c.stringWidth(nombre_eng, "Helvetica", 7)
        c.drawString(nombre_x_start + (nombre_width - tw) / 2, yp - 0.24 * inch, nombre_eng)
    
    yp = yp - logo_size - 0.05 * inch
    c.setFont("Helvetica", 7)
    net_weight = f"Net Wt. {p.get('Net_Weight', '')}"
    tw = c.stringWidth(net_weight, "Helvetica", 7)
    c.drawString(nombre_x_start + (nombre_width - tw) / 2, yp, net_weight)
    yp -= 10
    
    c.setFont("Helvetica-Bold", 10)
    c.drawString(x + PADDING, yp, "Nutrition Facts")
    yp -= 10
    
    c.setFont("Helvetica", 6)
    c.drawString(x + PADDING, yp, f"{p.get('Servings_Per_Container', '')} servings per container")
    yp -= 7
    
    c.setFont("Helvetica-Bold", 6)
    c.drawString(x + PADDING, yp, f"Serving size    {p.get('Serving_Size', '')}")
    yp -= 2
    
    c.setLineWidth(1.5)
    c.line(x + PADDING, yp, x + cfg['width'] - PADDING, yp)
    yp -= 5
    
    c.setFont("Helvetica", 5)
    c.drawString(x + PADDING, yp, "Amount per serving")
    yp -= 7
    
    c.setFont("Helvetica-Bold", 9)
    c.drawString(x + PADDING, yp, "Calories")
    c.drawString(x + PADDING + 45, yp, str(p.get('Calories', '')))
    yp -= 2
    
    c.setLineWidth(2.5)
    c.line(x + PADDING, yp, x + cfg['width'] - PADDING, yp)
    yp -= 5
    
    c.setFont("Helvetica-Bold", 5)
    c.drawRightString(x + cfg['width'] - PADDING, yp, "% Daily Value*")
    yp -= 7
    
    c.setLineWidth(0.3)
    
    def dn(label, campo, bold=True, indent=0):
        nonlocal yp
        valor = p.get(campo, '')
        dv = calc_dv(campo, valor)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 5)
        c.drawString(x + PADDING + indent, yp, f"{label} {valor}")
        if dv:
            c.drawRightString(x + cfg['width'] - PADDING, yp, f"{dv}%")
        yp -= 1
        c.line(x + PADDING, yp, x + cfg['width'] - PADDING, yp)
        yp -= 6
    
    dn("Total Fat", "Total_Fat")
    dn("Saturated Fat", "Saturated_Fat", False, 7)
    dn("Trans Fat", "Trans_Fat", False, 7)
    dn("Cholesterol", "Cholesterol")
    dn("Sodium", "Sodium")
    dn("Total Carb.", "Total_Carbohydrate")
    dn("Dietary Fiber", "Dietary_Fiber", False, 7)
    dn("Total Sugars", "Total_Sugars", False, 7)
    dn("  Incl. Added Sugars", "Added_Sugars", False, 10)
    
    c.setFont("Helvetica-Bold", 5)
    pv = p.get('Protein', '')
    pd = calc_dv('Protein', pv)
    c.drawString(x + PADDING, yp, f"Protein {pv}")
    if pd:
        c.drawRightString(x + cfg['width'] - PADDING, yp, f"{pd}%")
    yp -= 1
    
    c.setLineWidth(2.5)
    c.line(x + PADDING, yp, x + cfg['width'] - PADDING, yp)
    c.setLineWidth(0.3)
    yp -= 5
    
    c.setFont("Helvetica", 5)
    for campo, label in [('Vitamin_D', 'Vitamin D'), ('Calcium', 'Calcium'), 
                          ('Iron', 'Iron'), ('Potassium', 'Potassium')]:
        val = p.get(campo, '')
        if val and str(val).strip() != 'None':
            dv = calc_dv(campo, val)
            c.drawString(x + PADDING, yp, f"{label} {val}")
            if dv:
                c.drawRightString(x + cfg['width'] - PADDING, yp, f"{dv}%")
            yp -= 6
    
    c.setFont("Helvetica", 3.5)
    texto = "* The % Daily Value tells you how much a nutrient in a serving of food contributes to a daily diet. 2000 calories a day is used for general nutrition advice."
    max_width = cfg['width'] - 2 * PADDING
    linea = ""
    for palabra in texto.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 3.5) < max_width:
            linea = test
        else:
            c.drawString(x + PADDING, yp, linea.strip())
            yp -= 4.5
            linea = palabra + " "
    if linea:
        c.drawString(x + PADDING, yp, linea.strip())
        yp -= 5
    
    c.setFont("Helvetica-Bold", 5)
    c.drawString(x + PADDING, yp, "INGREDIENTS:")
    yp -= 5
    
    c.setFont("Helvetica", 4.5)
    ingredientes = str(p.get('Ingredients', ''))
    linea = ""
    for palabra in ingredientes.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 4.5) < max_width:
            linea = test
        else:
            c.drawString(x + PADDING, yp, linea.strip())
            yp -= 5
            linea = palabra + " "
    if linea:
        c.drawString(x + PADDING, yp, linea.strip())
        yp -= 6
    
    c.setFont("Helvetica-Bold", 5)
    allergens = p.get('Allergens', '')
    if allergens and str(allergens) != "NONE":
        c.drawString(x + PADDING, yp, allergens)
        yp -= 6
    
    c.setFont("Helvetica-Bold", 5.5)
    c.drawString(x + PADDING, yp, f"EXP: {p.get('Expiration_Date', '')}")
    yp -= 8
    
    c.setFont("Helvetica", 4.5)
    
    imported_by = p.get('Imported_By', '')
    has_imported = imported_by and str(imported_by).strip() and str(imported_by).strip() != 'None'
    
    if has_imported:
        ancho_total = cfg['width'] - 2 * PADDING
        ancho_cuadrito = ancho_total / 2 - 5
        
        left_x = x + PADDING
        c.drawString(left_x, yp, "Distributed by:")
        c.drawString(left_x, yp - 5, "Amaya Express")
        c.drawString(left_x, yp - 10, "472 Somerset St.,")
        c.drawString(left_x, yp - 15, "N. Plainfield, NJ 07060")
        
        right_x = x + PADDING + ancho_cuadrito + 10
        imported_parts = str(imported_by).split(',')
        
        c.drawString(right_x, yp, "Imported by:")
        if len(imported_parts) >= 1:
            c.drawString(right_x, yp - 5, imported_parts[0].strip())
        if len(imported_parts) >= 2:
            c.drawString(right_x, yp - 10, imported_parts[1].strip())
        if len(imported_parts) >= 3:
            c.drawString(right_x, yp - 15, imported_parts[2].strip())
    else:
        c.drawString(x + PADDING, yp, "Distributed by: Amaya Express")
        c.drawString(x + PADDING, yp - 5, "472 Somerset St., N. Plainfield, NJ 07060")


def dibujar_pls(c, x, y, p, cfg):
    """Dibujar etiqueta PLS504 (pequeña) con Nutrition Facts COMPLETA"""
    
    # Dimensiones
    etiqueta_ancho = cfg['width']
    etiqueta_alto = cfg['height']
    
    # División: Izquierda (40%) + Derecha (60% para Nutrition Facts)
    ancho_izquierda = etiqueta_ancho * 0.40
    ancho_derecha = etiqueta_ancho * 0.60
    x_derecha = x + ancho_izquierda
    
    # ========================================================================
    # LADO IZQUIERDO
    # ========================================================================
    yp = y + etiqueta_alto - PADDING * 0.2
    
    c.setFillColor(AMAYA_BLUE)
    c.setStrokeColor(black)
    
    # Logo Amaya
    logo_size = 0.25 * inch
    if os.path.exists(LOGO_PATH):
        c.drawImage(LOGO_PATH, x + PADDING * 0.2, yp - logo_size, 
                   logo_size, logo_size, preserveAspectRatio=True, mask='auto')
    
    # Product Name (al lado del logo)
    c.setFont("Helvetica-Bold", 6)
    nombre = p.get('Product_Name', '')
    nombre_x = x + PADDING * 0.2 + logo_size + 0.03 * inch
    
    # Dividir nombre si es muy largo
    max_width_nombre = ancho_izquierda - logo_size - PADDING * 0.4
    if c.stringWidth(nombre, "Helvetica-Bold", 6) > max_width_nombre:
        # Partir en dos líneas
        palabras = nombre.split()
        linea1 = ""
        linea2 = ""
        for palabra in palabras:
            test = linea1 + palabra + " "
            if c.stringWidth(test, "Helvetica-Bold", 6) < max_width_nombre:
                linea1 = test
            else:
                linea2 += palabra + " "
        c.drawString(nombre_x, yp - 0.06 * inch, linea1.strip())
        c.drawString(nombre_x, yp - 0.12 * inch, linea2.strip())
        nombre_y_offset = 0.18 * inch
    else:
        c.drawString(nombre_x, yp - 0.08 * inch, nombre)
        nombre_y_offset = 0.12 * inch
    
    # Product Name English
    nombre_eng = p.get('Product_Name_English', '')
    if nombre_eng and str(nombre_eng).strip() != 'None':
        c.setFont("Helvetica", 5)
        if c.stringWidth(nombre_eng, "Helvetica", 5) > max_width_nombre:
            nombre_eng = nombre_eng[:25] + "..."
        c.drawString(nombre_x, yp - nombre_y_offset, nombre_eng)
        nombre_y_offset += 0.06 * inch
    
    yp = yp - logo_size - 0.05 * inch
    
    # Net Weight
    c.setFont("Helvetica-Bold", 4.5)
    c.drawString(x + PADDING * 0.2, yp, f"Net Wt: {p.get('Net_Weight', '')}")
    yp -= 5
    
    # Expiration Date
    c.drawString(x + PADDING * 0.2, yp, f"EXP: {p.get('Expiration_Date', '')}")
    yp -= 6
    
    # Ingredients
    ingredients = p.get('Ingredients', '')
    if ingredients and str(ingredients).strip() != 'None':
        c.setFont("Helvetica-Bold", 3.5)
        c.drawString(x + PADDING * 0.2, yp, "Ingredients:")
        yp -= 3.5
        
        c.setFont("Helvetica", 3.5)
        # Dividir ingredientes en múltiples líneas
        max_width_ing = ancho_izquierda - PADDING * 0.4
        palabras = str(ingredients).split()
        linea = ""
        for palabra in palabras:
            test = linea + palabra + " "
            if c.stringWidth(test, "Helvetica", 3.5) < max_width_ing:
                linea = test
            else:
                c.drawString(x + PADDING * 0.2, yp, linea.strip())
                yp -= 3.5
                linea = palabra + " "
        if linea:
            c.drawString(x + PADDING * 0.2, yp, linea.strip())
            yp -= 3.5
        yp -= 2  # Espacio extra después de ingredientes
    
    # Allergens
    allergens = p.get('Allergens', '')
    if allergens and str(allergens).strip() != 'None':
        c.setFont("Helvetica-Bold", 3.5)
        max_width_allergens = ancho_izquierda - PADDING * 0.4
        if c.stringWidth(str(allergens), "Helvetica-Bold", 3.5) < max_width_allergens:
            c.drawString(x + PADDING * 0.2, yp, str(allergens))
        else:
            # Dividir si es muy largo
            allergens_short = str(allergens)[:30] + "..."
            c.drawString(x + PADDING * 0.2, yp, allergens_short)
        yp -= 5
    
    # Distributed by
    c.setFont("Helvetica-Bold", 3.5)
    c.drawString(x + PADDING * 0.2, yp, "Distributed by:")
    yp -= 3.5
    c.setFont("Helvetica", 3.2)
    c.drawString(x + PADDING * 0.2, yp, "Amaya Express Int'l")
    yp -= 3.2
    c.drawString(x + PADDING * 0.2, yp, "472 Somerset St.")
    yp -= 3.2
    c.drawString(x + PADDING * 0.2, yp, "N. Plainfield, NJ 06070")
    yp -= 5
    
    # Imported By
    imported_by = p.get('Imported_By', '')
    if imported_by and str(imported_by).strip() != 'None':
        c.setFont("Helvetica-Bold", 3.5)
        c.drawString(x + PADDING * 0.2, yp, "Imported by:")
        yp -= 3.5
        
        imported_parts = str(imported_by).split(',')
        c.setFont("Helvetica", 3.2)
        for part in imported_parts[:3]:  # Máximo 3 líneas
            part_text = part.strip()
            if len(part_text) > 28:
                part_text = part_text[:25] + "..."
            c.drawString(x + PADDING * 0.2, yp, part_text)
            yp -= 3.2
        
        # Bandera de El Salvador debajo de Imported By
        yp -= 2
        flag_width = 0.35 * inch
        flag_height = 0.2 * inch
        if os.path.exists(FLAG_PATH):
            c.drawImage(FLAG_PATH, x + PADDING * 0.2, yp - flag_height, 
                       flag_width, flag_height, preserveAspectRatio=True, mask='auto')
    
    # Información de contacto (abajo)
    contact_y = y + PADDING * 0.15 + 3.5
    c.setFont("Helvetica", 2.8)
    c.drawString(x + PADDING * 0.2, contact_y + 6, "(908) 405-5553")
    c.drawString(x + PADDING * 0.2, contact_y + 2.5, "amayaexpress21@hotmail.com")
    
    # ========================================================================
    # LADO DERECHO - NUTRITION FACTS COMPLETA
    # ========================================================================
    
    nf_x = x_derecha + PADDING * 0.1
    nf_y = y + PADDING * 0.25  # Más margen abajo
    nf_width = ancho_derecha - PADDING * 0.2
    nf_height = etiqueta_alto - PADDING * 0.45  # Cuadro más bajo
    
    # Borde de Nutrition Facts
    c.setFillColor(HexColor('#FFFFFF'))
    c.setStrokeColor(black)
    c.setLineWidth(1.2)
    c.rect(nf_x, nf_y, nf_width, nf_height, fill=1, stroke=1)
    
    c.setFillColor(AMAYA_BLUE)  # Texto azul
    c.setLineWidth(0.3)
    
    yp_nf = nf_y + nf_height - 3
    
    # Título "Nutrition Facts"
    c.setFont("Helvetica-Bold", 4.5)
    c.drawString(nf_x + 2, yp_nf, "Nutrition Facts")
    yp_nf -= 1
    c.setLineWidth(0.5)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 3
    
    # Servings
    c.setFont("Helvetica", 3)
    servings = p.get('Servings_Per_Container', '')
    if servings and str(servings).strip() != 'None':
        c.drawString(nf_x + 2, yp_nf, f"Serv: {servings}")
        yp_nf -= 3
    
    # Serving Size
    c.setFont("Helvetica-Bold", 3)
    serving_size = p.get('Serving_Size', '')
    ss_text = f"Size: {serving_size}"
    if c.stringWidth(ss_text, "Helvetica-Bold", 3) > nf_width - 4:
        ss_text = f"{serving_size}"
    c.drawString(nf_x + 2, yp_nf, ss_text)
    yp_nf -= 1
    
    c.setLineWidth(0.8)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 3
    
    # Calories
    c.setFont("Helvetica-Bold", 4)
    calories = p.get('Calories', '')
    c.drawString(nf_x + 2, yp_nf, "Calories")
    c.drawString(nf_x + 15, yp_nf, str(calories))
    yp_nf -= 1
    
    c.setLineWidth(1.2)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 2.5
    
    # % Daily Value
    c.setFont("Helvetica-Bold", 2.5)
    c.drawRightString(nf_x + nf_width - 2, yp_nf, "% DV*")
    yp_nf -= 1
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    yp_nf -= 3.2
    
    # Función para dibujar nutriente compacto
    def dn_compacto(label, campo, bold=True, indent=0):
        nonlocal yp_nf
        valor = p.get(campo, '')
        dv = calc_dv(campo, valor)
        
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 3.2)  # Era 2.8
        
        # Acortar labels
        label_map = {
            'Total Fat': 'Fat',
            'Saturated Fat': 'Sat Fat',
            'Trans Fat': 'Trans',
            'Total Carbohydrate': 'Carbs',
            'Dietary Fiber': 'Fiber',
            'Total Sugars': 'Sugars',
            'Added Sugars': 'Add Sug'
        }
        display_label = label_map.get(label, label)
        
        c.drawString(nf_x + 2 + indent, yp_nf, f"{display_label} {valor}")
        if dv:
            c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{dv}%")
        
        yp_nf -= 1.0  # Era 0.8
        c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
        yp_nf -= 3.2  # Era 2.8
    
    # Nutrientes principales
    dn_compacto("Total Fat", "Total_Fat")
    dn_compacto("Saturated Fat", "Saturated_Fat", False, 3)
    dn_compacto("Trans Fat", "Trans_Fat", False, 3)
    dn_compacto("Cholesterol", "Cholesterol")
    dn_compacto("Sodium", "Sodium")
    dn_compacto("Total Carbohydrate", "Total_Carbohydrate")
    dn_compacto("Dietary Fiber", "Dietary_Fiber", False, 3)
    dn_compacto("Total Sugars", "Total_Sugars", False, 3)
    dn_compacto("Added Sugars", "Added_Sugars", False, 5)
    
    # Protein
    c.setFont("Helvetica-Bold", 3.2)  # Era 2.8
    pv = p.get('Protein', '')
    pd = calc_dv('Protein', pv)
    c.drawString(nf_x + 2, yp_nf, f"Protein {pv}")
    if pd:
        c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{pd}%")
    yp_nf -= 1.0
    
    c.setLineWidth(1.2)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 3.2
    
    # Micronutrientes
    c.setFont("Helvetica", 3.2)  # Era 2.8
    for campo, label in [('Vitamin_D', 'Vit D'), ('Calcium', 'Calcium'), 
                          ('Iron', 'Iron'), ('Potassium', 'Potas.')]:
        val = p.get(campo, '')
        if val and str(val).strip() != 'None':
            dv = calc_dv(campo, val)
            c.drawString(nf_x + 2, yp_nf, f"{label} {val}")
            if dv:
                c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{dv}%")
            yp_nf -= 3.2  # Era 2.8
    
    # Leyenda del asterisco (muy pequeña)
    yp_nf -= 1
    c.setFont("Helvetica", 2)
    texto = "* The % Daily Value tells you how much a nutrient contributes to a daily diet. 2000 cal a day is used for general nutrition advice."
    max_width = nf_width - 4
    linea = ""
    for palabra in texto.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 2) < max_width:
            linea = test
        else:
            c.drawString(nf_x + 2, yp_nf, linea.strip())
            yp_nf -= 2.2
            linea = palabra + " "
    if linea:
        c.drawString(nf_x + 2, yp_nf, linea.strip())




    def dn_compacto(label, campo, bold=True, indent=0):
        nonlocal yp_nf
        valor = p.get(campo, '')
        dv = calc_dv(campo, valor)
        
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 2.8)
        
        # Acortar labels
        label_map = {
            'Total Fat': 'Fat',
            'Saturated Fat': 'Sat Fat',
            'Trans Fat': 'Trans',
            'Total Carbohydrate': 'Carbs',
            'Dietary Fiber': 'Fiber',
            'Total Sugars': 'Sugars',
            'Added Sugars': 'Add Sug'
        }
        display_label = label_map.get(label, label)
        
        c.drawString(nf_x + 2 + indent, yp_nf, f"{display_label} {valor}")
        if dv:
            c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{dv}%")
        
        yp_nf -= 0.8
        c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
        yp_nf -= 2.8
    
    # Nutrientes principales
    dn_compacto("Total Fat", "Total_Fat")
    dn_compacto("Saturated Fat", "Saturated_Fat", False, 3)
    dn_compacto("Trans Fat", "Trans_Fat", False, 3)
    dn_compacto("Cholesterol", "Cholesterol")
    dn_compacto("Sodium", "Sodium")
    dn_compacto("Total Carbohydrate", "Total_Carbohydrate")
    dn_compacto("Dietary Fiber", "Dietary_Fiber", False, 3)
    dn_compacto("Total Sugars", "Total_Sugars", False, 3)
    dn_compacto("Added Sugars", "Added_Sugars", False, 5)
    
    # Protein
    c.setFont("Helvetica-Bold", 2.8)
    pv = p.get('Protein', '')
    pd = calc_dv('Protein', pv)
    c.drawString(nf_x + 2, yp_nf, f"Protein {pv}")
    if pd:
        c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{dv}%")
    yp_nf -= 0.8
    
    c.setLineWidth(1.2)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 2.8
    
    # Micronutrientes
    c.setFont("Helvetica", 2.8)
    for campo, label in [('Vitamin_D', 'Vit D'), ('Calcium', 'Calcium'), 
                          ('Iron', 'Iron'), ('Potassium', 'Potas.')]:
        val = p.get(campo, '')
        if val and str(val).strip() != 'None':
            dv = calc_dv(campo, val)
            c.drawString(nf_x + 2, yp_nf, f"{label} {val}")
            if dv:
                c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{dv}%")
            yp_nf -= 2.8
    
    # Leyenda del asterisco (muy pequeña)
    yp_nf -= 1
    c.setFont("Helvetica", 2)
    texto = "* The % Daily Value tells you how much a nutrient contributes to a daily diet. 2000 cal a day is used for general nutrition advice."
    max_width = nf_width - 4
    linea = ""
    for palabra in texto.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 2) < max_width:
            linea = test
        else:
            c.drawString(nf_x + 2, yp_nf, linea.strip())
            yp_nf -= 2.2
            linea = palabra + " "
    if linea:
        c.drawString(nf_x + 2, yp_nf, linea.strip())


def dibujar_lacteo_san_julian(c, x, y, p, cfg):
    """
    Dibujar etiqueta San Julián - Diseño vertical rotado 90°
    Basado en Lacteo Avery 8164 pero con diseño San Julián
    
    IMPORTANTE: La etiqueta se diseña vertical (3.33" x 4") pero se rota 90°
    para usar el espacio horizontal de Lacteo Avery 8164 (4" x 3.33")
    """
    
    # ROTAR EL CANVAS 90 GRADOS
    # Guardamos el estado actual del canvas
    c.saveState()
    
    # La etiqueta vertical real será: ancho_real=3.33", alto_real=4"
    ancho_real = 3.33 * inch
    alto_real = 4 * inch
    
    # Trasladar y rotar para que la etiqueta vertical ocupe el espacio horizontal
    c.translate(x + cfg['width'], y)
    c.rotate(90)
    
    # Ahora trabajamos en un espacio de 3.33" (ancho) × 4" (alto)
    # Las coordenadas (0, 0) están en la esquina inferior izquierda de la etiqueta rotada
    
    # ========================================================================
    # FONDO DEGRADADO AZUL
    # ========================================================================
    if os.path.exists(FONDO_AZUL_SAN_JULIAN_PATH):
        c.drawImage(FONDO_AZUL_SAN_JULIAN_PATH, 
                    0, 0,  # Esquina inferior izquierda
                    width=ancho_real, 
                    height=alto_real,
                    preserveAspectRatio=False, 
                    mask='auto')
    
    # ========================================================================
    # CONTENIDO DE LA ETIQUETA (de arriba hacia abajo)
    # ========================================================================
    
    yp = alto_real - PADDING * 0.5  # Empezar desde arriba
    
    # ------------------------------------------------------------------------
    # 1. TEXTOS SUPERIORES: "Keep refrigerated / Pasteurizado"
    # ------------------------------------------------------------------------
    c.setFillColor(HexColor('#FFFFFF'))  # Texto blanco sobre azul
    c.setFont("Helvetica-Bold", 7)
    
    keep_text = "Keep refrigerated"
    text_width = c.stringWidth(keep_text, "Helvetica-Bold", 7)
    c.drawString((ancho_real - text_width) / 2, yp, keep_text)
    yp -= 8
    
    pasterizado_text = "Pasteurizado"
    text_width = c.stringWidth(pasterizado_text, "Helvetica-Bold", 7)
    c.drawString((ancho_real - text_width) / 2, yp, pasterizado_text)
    yp -= 15
    
    # ------------------------------------------------------------------------
    # 2. LOGO SAN JULIÁN (Vaca con nombre)
    # ------------------------------------------------------------------------
    if os.path.exists(LOGO_SAN_JULIAN_PATH):
        logo_size = 0.6 * inch
        logo_x = (ancho_real - logo_size) / 2
        c.drawImage(LOGO_SAN_JULIAN_PATH, 
                    logo_x, yp - logo_size,
                    logo_size, logo_size, 
                    preserveAspectRatio=True, 
                    mask='auto')
        yp -= logo_size + 0.1 * inch
    
    # ------------------------------------------------------------------------
    # 3. NOMBRE DEL PRODUCTO (español - grande)
    # ------------------------------------------------------------------------
    c.setFillColor(HexColor('#FFFFFF'))  # Blanco
    nombre_producto = p.get('Product_Name', '')
    c.setFont("Helvetica-Bold", 16)
    
    # Si es muy largo, dividir en dos líneas
    max_width_nombre = ancho_real - PADDING
    if c.stringWidth(nombre_producto, "Helvetica-Bold", 16) > max_width_nombre:
        palabras = nombre_producto.split()
        linea1 = ""
        linea2 = ""
        for palabra in palabras:
            test = linea1 + palabra + " "
            if c.stringWidth(test, "Helvetica-Bold", 16) < max_width_nombre:
                linea1 = test
            else:
                linea2 += palabra + " "
        
        # Dibujar línea 1
        text_width = c.stringWidth(linea1.strip(), "Helvetica-Bold", 16)
        c.drawString((ancho_real - text_width) / 2, yp, linea1.strip())
        yp -= 18
        
        # Dibujar línea 2
        if linea2:
            text_width = c.stringWidth(linea2.strip(), "Helvetica-Bold", 16)
            c.drawString((ancho_real - text_width) / 2, yp, linea2.strip())
            yp -= 18
    else:
        text_width = c.stringWidth(nombre_producto, "Helvetica-Bold", 16)
        c.drawString((ancho_real - text_width) / 2, yp, nombre_producto)
        yp -= 18
    
    # ------------------------------------------------------------------------
    # 4. NOMBRE EN INGLÉS (Product_Name_English - cursiva)
    # ------------------------------------------------------------------------
    nombre_eng = p.get('Product_Name_English', '')
    if nombre_eng and str(nombre_eng).strip() != 'None':
        c.setFont("Helvetica-Oblique", 11)
        text_width = c.stringWidth(nombre_eng, "Helvetica-Oblique", 11)
        c.drawString((ancho_real - text_width) / 2, yp, nombre_eng)
        yp -= 15
    
    yp -= 10  # Espacio adicional antes de la sección inferior
    
    # ========================================================================
    # SECCIÓN INFERIOR: Dividida en dos columnas
    # Izquierda: Sello + Ingredientes + Allergens + Peso
    # Derecha: Nutrition Facts COMPLETA
    # ========================================================================
    
    # Definir anchos de columnas
    ancho_izquierda = ancho_real * 0.45
    ancho_derecha = ancho_real * 0.55
    x_derecha = ancho_izquierda
    
    # Altura disponible para la sección inferior
    altura_inferior = yp - PADDING * 0.5
    y_inicio_inferior = PADDING * 0.5
    
    # ------------------------------------------------------------------------
    # COLUMNA IZQUIERDA
    # ------------------------------------------------------------------------
    yp_izq = altura_inferior + y_inicio_inferior
    
    # SELLO VERDE "Sabor San Julián"
    if os.path.exists(SELLO_SAN_JULIAN_PATH):
        sello_size = 0.45 * inch
        sello_x = PADDING * 0.3
        c.drawImage(SELLO_SAN_JULIAN_PATH, 
                    sello_x, yp_izq - sello_size,
                    sello_size, sello_size, 
                    preserveAspectRatio=True, 
                    mask='auto')
        yp_izq -= sello_size + 0.05 * inch
    
    # INGREDIENTES
    c.setFillColor(HexColor('#FFFFFF'))
    c.setFont("Helvetica-Bold", 4.5)
    c.drawString(PADDING * 0.3, yp_izq, "INGREDIENTS:")
    yp_izq -= 5
    
    c.setFont("Helvetica", 4)
    ingredientes = str(p.get('Ingredients', ''))
    max_width_ing = ancho_izquierda - PADDING * 0.6
    linea = ""
    for palabra in ingredientes.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 4) < max_width_ing:
            linea = test
        else:
            c.drawString(PADDING * 0.3, yp_izq, linea.strip())
            yp_izq -= 4.5
            linea = palabra + " "
    if linea:
        c.drawString(PADDING * 0.3, yp_izq, linea.strip())
    yp_izq -= 7
    
    # ALLERGENS
    allergens = p.get('Allergens', '')
    if allergens and str(allergens).strip() != 'None':
        c.setFont("Helvetica-Bold", 4.5)
        c.drawString(PADDING * 0.3, yp_izq, str(allergens))
        yp_izq -= 7
    
    # PESO NETO
    c.setFont("Helvetica-Bold", 5)
    peso_text = f"Net Wt: {p.get('Net_Weight', '')}"
    c.drawString(PADDING * 0.3, yp_izq, peso_text)
    yp_izq -= 7
    
    # PRODUCT OF EL SALVADOR
    c.setFont("Helvetica", 4.5)
    c.drawString(PADDING * 0.3, yp_izq, "Product of El Salvador")
    yp_izq -= 10  # 2 líneas de espacio
    
    # FECHA DE VENCIMIENTO (solo fecha, sin hora)
    c.setFont("Helvetica-Bold", 5)
    exp_date = str(p.get('Expiration_Date', ''))
    # Si la fecha tiene hora (formato con espacio), tomar solo la parte de la fecha
    if ' ' in exp_date:
        exp_date = exp_date.split(' ')[0]
    exp_text = f"EXP: {exp_date}"
    c.drawString(PADDING * 0.3, yp_izq, exp_text)
    
    # ------------------------------------------------------------------------
    # COLUMNA DERECHA: NUTRITION FACTS COMPLETA
    # ------------------------------------------------------------------------
    nf_x = x_derecha + PADDING * 0.2
    nf_y = y_inicio_inferior + 0.1 * inch
    nf_width = ancho_derecha - PADDING * 0.4
    nf_height = altura_inferior - 0.2 * inch
    
    # Borde de Nutrition Facts (fondo blanco)
    c.setFillColor(HexColor('#FFFFFF'))
    c.setStrokeColor(black)
    c.setLineWidth(1.5)
    c.rect(nf_x, nf_y, nf_width, nf_height, fill=1, stroke=1)
    
    # Contenido de Nutrition Facts
    c.setFillColor(black)
    c.setLineWidth(0.3)
    yp_nf = nf_y + nf_height - 4
    
    # Título
    c.setFont("Helvetica-Bold", 6)
    c.drawString(nf_x + 2, yp_nf, "Nutrition Facts")
    yp_nf -= 1
    c.setLineWidth(0.8)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 4
    
    # Servings
    c.setFont("Helvetica", 4)
    servings = p.get('Servings_Per_Container', '')
    if servings and str(servings).strip() != 'None':
        c.drawString(nf_x + 2, yp_nf, f"Servings: {servings}")
        yp_nf -= 4
    
    # Serving Size
    c.setFont("Helvetica-Bold", 4)
    serving_size = p.get('Serving_Size', '')
    c.drawString(nf_x + 2, yp_nf, f"Serving size")
    c.drawRightString(nf_x + nf_width - 2, yp_nf, str(serving_size))
    yp_nf -= 1
    c.setLineWidth(1.2)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 4
    
    # Calories
    c.setFont("Helvetica-Bold", 3.5)
    c.drawString(nf_x + 2, yp_nf, "Calories")
    c.setFont("Helvetica-Bold", 6)
    calories = p.get('Calories', '')
    c.drawRightString(nf_x + nf_width - 2, yp_nf, str(calories))
    yp_nf -= 1
    c.setLineWidth(1)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 3
    
    # % Daily Value
    c.setFont("Helvetica-Bold", 3)
    c.drawRightString(nf_x + nf_width - 2, yp_nf, "% DV*")
    yp_nf -= 1
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    yp_nf -= 4
    
    # Función auxiliar para dibujar nutrientes
    def dn(label, campo, bold=True, indent=0):
        nonlocal yp_nf
        valor = p.get(campo, '')
        dv = calc_dv(campo, valor)
        c.setFont("Helvetica-Bold" if bold else "Helvetica", 4)
        c.drawString(nf_x + 2 + indent, yp_nf, f"{label} {valor}")
        if dv:
            c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{dv}%")
        yp_nf -= 1
        c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
        yp_nf -= 4
    
    # Nutrientes completos
    dn("Total Fat", "Total_Fat")
    dn("Saturated Fat", "Saturated_Fat", False, 5)
    dn("Trans Fat", "Trans_Fat", False, 5)
    dn("Cholesterol", "Cholesterol")
    dn("Sodium", "Sodium")
    dn("Total Carbohydrate", "Total_Carbohydrate")
    dn("Dietary Fiber", "Dietary_Fiber", False, 5)
    dn("Total Sugars", "Total_Sugars", False, 5)
    dn(" Incl. Added Sugars", "Added_Sugars", False, 8)
    
    # Protein
    c.setFont("Helvetica-Bold", 4)
    pv = p.get('Protein', '')
    pd = calc_dv('Protein', pv)
    c.drawString(nf_x + 2, yp_nf, f"Protein {pv}")
    if pd:
        c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{pd}%")
    yp_nf -= 1
    c.setLineWidth(1.2)
    c.line(nf_x + 2, yp_nf, nf_x + nf_width - 2, yp_nf)
    c.setLineWidth(0.3)
    yp_nf -= 4
    
    # Micronutrientes
    c.setFont("Helvetica", 4)
    for campo, label in [('Vitamin_D', 'Vitamin D'), ('Calcium', 'Calcium'), 
                          ('Iron', 'Iron'), ('Potassium', 'Potassium')]:
        val = p.get(campo, '')
        if val and str(val).strip() != 'None':
            dv = calc_dv(campo, val)
            c.drawString(nf_x + 2, yp_nf, f"{label} {val}")
            if dv:
                c.drawRightString(nf_x + nf_width - 2, yp_nf, f"{dv}%")
            yp_nf -= 4
    
    # Leyenda del asterisco
    yp_nf -= 1
    c.setFont("Helvetica", 2.5)
    texto = "* The % Daily Value tells you how much a nutrient in a serving contributes to a daily diet. 2000 cal a day is used for general nutrition advice."
    max_width = nf_width - 4
    linea = ""
    for palabra in texto.split():
        test = linea + palabra + " "
        if c.stringWidth(test, "Helvetica", 2.5) < max_width:
            linea = test
        else:
            c.drawString(nf_x + 2, yp_nf, linea.strip())
            yp_nf -= 3
            linea = palabra + " "
    if linea:
        c.drawString(nf_x + 2, yp_nf, linea.strip())
    
    # RESTAURAR EL ESTADO DEL CANVAS (deshacer la rotación)
    c.restoreState()


def wrap_texto(c, texto, fuente, tam, ancho_max, xi, ypos, interlinea, dibujar=True):
    """
    Dibuja texto con salto de línea automático. Devuelve la posición Y final.
    Con dibujar=False no escribe nada (solo calcula el alto que ocuparía) - se usa
    para medir el bloque de texto antes de centrarlo verticalmente (ver AVERY_5260).
    """
    c.setFont(fuente, tam)
    linea = ""
    for palabra in str(texto).split():
        test = linea + palabra + " "
        if c.stringWidth(test, fuente, tam) < ancho_max:
            linea = test
        else:
            if dibujar:
                c.drawString(xi, ypos, linea.strip())
            ypos -= interlinea
            linea = palabra + " "
    if linea:
        if dibujar:
            c.drawString(xi, ypos, linea.strip())
        ypos -= interlinea
    return ypos


def dibujar_avery5260(c, x, y, p, cfg):
    """
    AVERY 5260 ("Dulces") - 1" x 2-5/8" - Etiqueta de identificación + Nutrition
    Facts compacta. Migrada desde el prototipo probado e impreso en papel real
    en Cowork (agosto 2026), empezando con el producto Coco Rallado.

    Esta etiqueta NO tiene que cumplir la restricción de tamaño mínimo de fuente
    de la FDA (21 CFR 101.9 exige 14pt mínimo en Nutrition Facts) porque no se usa
    para ese fin - por eso el texto va en 3.6-6.5pt.

    Layout: columna izquierda (nombre, INGREDIENTS, Distributed by fijo +
    Imported_By del Excel, Net Wt/EXP, teléfonos/correo) + columna derecha con
    Nutrition Facts, separadas por 2 líneas verticales delgadas (0.4pt) + 1 línea
    horizontal arriba cerrando el cuadro. Ninguna línea llega al borde físico de
    la etiqueta (para que no se corte al imprimir/recortar).

    Los nutrientes en 0 (o vacíos) se omiten (ver es_cero()); las filas que sí
    quedan se reparten para ocupar todo el alto de la columna de Nutrition Facts,
    reduciendo el tamaño de letra automáticamente (mínimo 2.6pt) solo si hacen
    falta muchas filas. La columna izquierda va centrada verticalmente (se mide
    el bloque de texto antes de dibujarlo) para que no se corte al despegar la
    etiqueta - la columna de Nutrition Facts se probó centrada igual pero se
    descartó porque obligaba a achicar más la letra, así que se dejó a todo el
    alto de la etiqueta.
    """
    c.setFillColor(black)
    c.setStrokeColor(black)

    nf_width = 0.90 * inch
    gap = 0.05 * inch
    right_margin = 0.07 * inch
    nf_x = x + cfg['width'] - right_margin - nf_width
    left_x = x + 0.10 * inch
    left_width = cfg['width'] - nf_width - gap - 0.08 * inch

    # ---------- COLUMNA IZQUIERDA (centrada verticalmente) ----------
    def dibujar_columna_izquierda(yp_inicial, dibujar=True):
        yp = yp_inicial

        nombre = str(p.get('Product_Name', '') or '').strip()
        if dibujar:
            c.setFont("Helvetica-Bold", 6.5)
            c.drawString(left_x, yp, nombre)
        yp -= 6.5

        nombre_eng = p.get('Product_Name_English', '')
        if nombre_eng and str(nombre_eng).strip() not in ('', 'None'):
            if dibujar:
                c.setFont("Helvetica-Oblique", 5)
                c.drawString(left_x, yp, str(nombre_eng).strip())
            yp -= 6

        ingredientes = p.get('Ingredients', '')
        if ingredientes and str(ingredientes).strip() not in ('', 'None'):
            if dibujar:
                c.setFont("Helvetica-Bold", 4.2)
                c.drawString(left_x, yp, "INGREDIENTS:")
            yp -= 4.5
            yp = wrap_texto(c, ingredientes, "Helvetica", 4, left_width, left_x, yp, 4.2, dibujar)

        yp -= 1
        if dibujar:
            c.setFont("Helvetica-Bold", 4)
            c.drawString(left_x, yp, "Distributed by:")
        yp -= 4
        if dibujar:
            c.setFont("Helvetica", 3.7)
            c.drawString(left_x, yp, "Amaya Express Int'l, 472 Somerset St,")
        yp -= 3.9
        if dibujar:
            c.drawString(left_x, yp, "North Plainfield, NJ 06070")
        yp -= 4.3

        imported_by = p.get('Imported_By', '')
        if imported_by and str(imported_by).strip() not in ('', 'None'):
            if dibujar:
                c.setFont("Helvetica-Bold", 4)
                c.drawString(left_x, yp, "Imported by:")
            yp -= 4
            if dibujar:
                c.setFont("Helvetica", 3.7)
            for parte in str(imported_by).split(',')[:2]:
                if dibujar:
                    c.drawString(left_x, yp, parte.strip())
                yp -= 3.9

        net_weight = p.get('Net_Weight', '')
        exp = p.get('Expiration_Date', '')
        pie = []
        if net_weight and str(net_weight).strip() not in ('', 'None'):
            pie.append(f"Net Wt: {net_weight}")
        if exp and str(exp).strip() not in ('', 'None'):
            pie.append(f"EXP: {formatear_exp(exp)}")
        if pie:
            yp -= 1
            if dibujar:
                c.setFont("Helvetica-Bold", 4.2)
                c.drawString(left_x, yp, "   ".join(pie))
            yp -= 4.5

        if dibujar:
            c.setFont("Helvetica", 3.6)
            c.drawString(left_x, yp, "(908) 405-5553 / (908) 405-3072")
        yp -= 3.8
        if dibujar:
            c.drawString(left_x, yp, "amayaexpress21@gmail.com")

        return yp

    # 1) Medir el alto real del bloque (sin dibujar nada)
    yp_final_medido = dibujar_columna_izquierda(0, dibujar=False)
    alto_columna_izquierda = 0 - yp_final_medido

    # 2) Margen simétrico arriba/abajo dentro de la etiqueta (mínimo 2pt de seguridad)
    margen_vertical = max((cfg['height'] - alto_columna_izquierda) / 2, 2)

    # 3) Dibujar de verdad, ya centrado
    dibujar_columna_izquierda(y + cfg['height'] - margen_vertical, dibujar=True)

    # ---------- COLUMNA NUTRITION FACTS (todo el alto de la etiqueta) ----------
    nf_top = y + cfg['height']
    nf_bottom = y
    linea_vertical_top = nf_top - 1.5  # el final se calcula después de dibujar las filas

    # Línea horizontal de arriba, cerrando el cuadro junto con las dos verticales
    c.setLineWidth(0.4)
    c.line(nf_x - gap / 2, linea_vertical_top,
           nf_x + nf_width + right_margin / 2, linea_vertical_top)

    yp_nf = nf_top - 6.5
    c.setFont("Helvetica-Bold", 6)
    c.drawString(nf_x, yp_nf, "Nutrition Facts")
    yp_nf -= 6

    serving_size = p.get('Serving_Size', '')
    if serving_size and str(serving_size).strip() not in ('', 'None'):
        c.setFont("Helvetica", 3.6)
        c.drawString(nf_x, yp_nf, f"Serving Size {serving_size}")
        yp_nf -= 4

    c.setLineWidth(1)
    c.line(nf_x, yp_nf, nf_x + nf_width, yp_nf)
    yp_nf -= 4.5

    calories = p.get('Calories', '')
    if calories and str(calories).strip() not in ('', 'None'):
        c.setFont("Helvetica-Bold", 6)
        c.drawString(nf_x, yp_nf, "Calories")
        c.drawString(nf_x + 32, yp_nf, str(calories))
        yp_nf -= 2
        c.setLineWidth(1.8)
        c.line(nf_x, yp_nf, nf_x + nf_width, yp_nf)
        yp_nf -= 4

    c.setFont("Helvetica-Bold", 3.3)
    c.drawRightString(nf_x + nf_width, yp_nf, "% DV*")
    yp_nf -= 3.5

    header_bottom = yp_nf  # a partir de aquí se reparten las filas de nutrientes

    filas = [
        ("Total Fat", "Total_Fat", True, 0),
        ("Sat. Fat", "Saturated_Fat", False, 4),
        ("Trans Fat", "Trans_Fat", False, 4),
        ("Cholesterol", "Cholesterol", True, 0),
        ("Sodium", "Sodium", True, 0),
        ("Total Carb.", "Total_Carbohydrate", True, 0),
        ("Fiber", "Dietary_Fiber", False, 4),
        ("Sugars", "Total_Sugars", False, 4),
        ("  Added Sugars", "Added_Sugars", False, 6),
        ("Protein", "Protein", True, 0),
        ("Vitamin D", "Vitamin_D", False, 0),
        ("Calcium", "Calcium", False, 0),
        ("Iron", "Iron", False, 0),
        ("Potassium", "Potassium", False, 0),
    ]
    filas_visibles = [f for f in filas if not es_cero(p.get(f[1], ''))]

    n = len(filas_visibles)
    if n:
        espacio_disponible = max(header_bottom - nf_bottom - 2, 0)

        # Cada fila necesita un mínimo de espacio para que el texto no choque con
        # la línea divisoria de abajo. Empezamos en 3.6pt y solo achicamos la letra
        # si con muchos nutrientes no alcanza el alto disponible; nunca bajamos de 2.6pt.
        tam_fuente = 3.6
        pre_linea = 0.9
        post_linea_min = tam_fuente
        while n * (pre_linea + post_linea_min) > espacio_disponible and tam_fuente > 2.6:
            tam_fuente -= 0.1
            post_linea_min = tam_fuente

        paso = espacio_disponible / n
        post_linea = max(post_linea_min, paso - pre_linea)

        y_fila = header_bottom
        for label, campo, bold, indent in filas_visibles:
            valor = p.get(campo, '')
            dv = calc_dv(campo, valor)
            c.setFont("Helvetica-Bold" if bold else "Helvetica", tam_fuente)
            c.drawString(nf_x + indent, y_fila, f"{label} {valor}")
            if dv:
                c.drawRightString(nf_x + nf_width, y_fila, f"{dv}%")
            y_fila -= pre_linea
            c.setLineWidth(0.3)
            c.line(nf_x, y_fila, nf_x + nf_width, y_fila)
            y_fila -= post_linea
        linea_vertical_bottom = y_fila + post_linea
    else:
        linea_vertical_bottom = header_bottom

    # Separadores verticales (izquierda y derecha), solo del alto del contenido real
    c.setLineWidth(0.4)
    c.line(nf_x - gap / 2, linea_vertical_bottom, nf_x - gap / 2, linea_vertical_top)
    c.line(nf_x + nf_width + right_margin / 2, linea_vertical_bottom,
           nf_x + nf_width + right_margin / 2, linea_vertical_top)


# ============================================================================
# CLASE PRINCIPAL - INTERFAZ GRÁFICA
# ============================================================================
class EtiquetasApp(QMainWindow):
    def __init__(self):
        super().__init__()
        self.productos = []
        self.checkboxes = []
        self.spinboxes = []
        self.initUI()
        self.cargar_productos()
    
    def initUI(self):
        """Crear la interfaz gráfica"""
        self.setWindowTitle(f"Generador de Etiquetas - Amaya Express")
        
        # Widget central
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        
        # Layout principal
        main_layout = QVBoxLayout()
        central_widget.setLayout(main_layout)
        
        # ====================================================================
        # HEADER CON LOGO
        # ====================================================================
        header_container = QWidget()
        header_layout = QHBoxLayout()
        header_container.setLayout(header_layout)
        header_container.setStyleSheet("""
            QWidget {
                background-color: #0066B3;
                border-radius: 5px;
            }
        """)
        
        # Logo Amaya (izquierda)
        if os.path.exists(LOGO_PATH):
            logo_label = QLabel()
            pixmap = QPixmap(LOGO_PATH)
            # Escalar el logo a altura de 80px manteniendo proporción
            scaled_pixmap = pixmap.scaledToHeight(80, Qt.SmoothTransformation)
            logo_label.setPixmap(scaled_pixmap)
            logo_label.setStyleSheet("background-color: transparent; padding: 10px;")
            header_layout.addWidget(logo_label)
        
        # Título (centro)
        header_title = QLabel("GENERADOR DE ETIQUETAS FDA")
        header_title.setStyleSheet("""
            QLabel {
                background-color: transparent;
                color: white;
                font-size: 24px;
                font-weight: bold;
                padding: 20px;
            }
        """)
        header_title.setAlignment(Qt.AlignCenter)
        header_layout.addWidget(header_title, stretch=1)
        
        # Versión (derecha)
        version_label = QLabel(f"v{VERSION}")
        version_label.setStyleSheet("""
            QLabel {
                background-color: transparent;
                color: rgba(255, 255, 255, 0.7);
                font-size: 11px;
                padding: 10px;
            }
        """)
        version_label.setAlignment(Qt.AlignRight | Qt.AlignTop)
        header_layout.addWidget(version_label)
        
        main_layout.addWidget(header_container)
        
        # ====================================================================
        # SELECTOR DE TIPO DE ETIQUETA
        # ====================================================================
        tipo_layout = QHBoxLayout()
        
        tipo_label = QLabel("Tipo de Etiqueta:")
        tipo_label.setFont(QFont("Arial", 12, QFont.Bold))
        tipo_layout.addWidget(tipo_label)
        
        self.tipo_combo = QComboBox()
        # IMPORTANTE: Agregar opción por defecto (sin selección)
        self.tipo_combo.addItem("-- Seleccione tipo de etiqueta --")
        self.tipo_combo.addItems([
            "Avery 8164 (6 por hoja - Vertical)",
            "Lacteo Avery 8164 (6 por hoja - Horizontal)",
            "Lacteo San Julian (6 por hoja - Vertical rotada)",
            "PLS 504 (10 por hoja - Pequeña)",
            "Dulces - Avery 5260 (30 por hoja)"
        ])
        self.tipo_combo.setCurrentIndex(0)  # Por defecto en la opción vacía
        self.tipo_combo.setFont(QFont("Arial", 11))
        self.tipo_combo.setStyleSheet("""
            QComboBox {
                padding: 8px;
                border: 2px solid #0066B3;
                border-radius: 5px;
                background: white;
            }
        """)
        self.tipo_combo.currentIndexChanged.connect(self.actualizar_resumen)
        tipo_layout.addWidget(self.tipo_combo)
        
        tipo_layout.addStretch()
        main_layout.addLayout(tipo_layout)
        
        # ====================================================================
        # BUSCADOR DE PRODUCTOS
        # ====================================================================
        buscador_layout = QHBoxLayout()
        
        buscador_label = QLabel("🔍 Buscar Producto:")
        buscador_label.setFont(QFont("Arial", 11, QFont.Bold))
        buscador_layout.addWidget(buscador_label)
        
        self.search_input = QLineEdit()
        self.search_input.setPlaceholderText("Escribe para buscar por nombre...")
        self.search_input.setFont(QFont("Arial", 11))
        self.search_input.setStyleSheet("""
            QLineEdit {
                padding: 10px;
                border: 2px solid #dee2e6;
                border-radius: 5px;
                background: white;
            }
            QLineEdit:focus {
                border-color: #0066B3;
            }
        """)
        self.search_input.textChanged.connect(self.filtrar_productos)
        buscador_layout.addWidget(self.search_input)
        
        main_layout.addLayout(buscador_layout)
        
        # ====================================================================
        # BOTONES DE SELECCIÓN
        # ====================================================================
        botones_layout = QHBoxLayout()
        
        btn_seleccionar_todo = QPushButton("✓ Seleccionar Todo")
        btn_seleccionar_todo.setFont(QFont("Arial", 10))
        btn_seleccionar_todo.setStyleSheet("""
            QPushButton {
                background-color: #28a745;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #218838;
            }
        """)
        btn_seleccionar_todo.clicked.connect(self.seleccionar_todo)
        botones_layout.addWidget(btn_seleccionar_todo)
        
        btn_deseleccionar_todo = QPushButton("✗ Deseleccionar Todo")
        btn_deseleccionar_todo.setFont(QFont("Arial", 10))
        btn_deseleccionar_todo.setStyleSheet("""
            QPushButton {
                background-color: #6c757d;
                color: white;
                padding: 10px 20px;
                border: none;
                border-radius: 5px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #5a6268;
            }
        """)
        btn_deseleccionar_todo.clicked.connect(self.deseleccionar_todo)
        botones_layout.addWidget(btn_deseleccionar_todo)
        
        botones_layout.addStretch()
        main_layout.addLayout(botones_layout)
        
        # ====================================================================
        # ÁREA DE PRODUCTOS (con scroll)
        # ====================================================================
        scroll_area = QScrollArea()
        scroll_area.setWidgetResizable(True)
        scroll_area.setStyleSheet("""
            QScrollArea {
                border: 2px solid #dee2e6;
                border-radius: 5px;
            }
        """)
        
        self.productos_widget = QWidget()
        self.productos_layout = QVBoxLayout()
        self.productos_widget.setLayout(self.productos_layout)
        
        scroll_area.setWidget(self.productos_widget)
        main_layout.addWidget(scroll_area)
        
        # ====================================================================
        # RESUMEN
        # ====================================================================
        resumen_group = QGroupBox("📊 Resumen")
        resumen_group.setFont(QFont("Arial", 12, QFont.Bold))
        resumen_group.setStyleSheet("""
            QGroupBox {
                border: 2px solid #0066B3;
                border-radius: 5px;
                margin-top: 10px;
                padding-top: 10px;
            }
            QGroupBox::title {
                color: #0066B3;
            }
        """)
        
        resumen_layout = QGridLayout()
        
        self.lbl_productos_sel = QLabel("0")
        self.lbl_productos_sel.setFont(QFont("Arial", 20, QFont.Bold))
        self.lbl_productos_sel.setStyleSheet("color: #0066B3;")
        self.lbl_productos_sel.setAlignment(Qt.AlignCenter)
        
        self.lbl_total_etiquetas = QLabel("0")
        self.lbl_total_etiquetas.setFont(QFont("Arial", 20, QFont.Bold))
        self.lbl_total_etiquetas.setStyleSheet("color: #28a745;")
        self.lbl_total_etiquetas.setAlignment(Qt.AlignCenter)
        
        self.lbl_hojas = QLabel("0")
        self.lbl_hojas.setFont(QFont("Arial", 20, QFont.Bold))
        self.lbl_hojas.setStyleSheet("color: #dc3545;")
        self.lbl_hojas.setAlignment(Qt.AlignCenter)
        
        resumen_layout.addWidget(QLabel("Productos Seleccionados:"), 0, 0)
        resumen_layout.addWidget(self.lbl_productos_sel, 1, 0)
        
        resumen_layout.addWidget(QLabel("Total de Etiquetas:"), 0, 1)
        resumen_layout.addWidget(self.lbl_total_etiquetas, 1, 1)
        
        resumen_layout.addWidget(QLabel("Hojas a Imprimir:"), 0, 2)
        resumen_layout.addWidget(self.lbl_hojas, 1, 2)
        
        resumen_group.setLayout(resumen_layout)
        main_layout.addWidget(resumen_group)
        
        # ====================================================================
        # BOTÓN GENERAR PDF
        # ====================================================================
        self.btn_generar = QPushButton("📄 GENERAR PDF")
        self.btn_generar.setFont(QFont("Arial", 16, QFont.Bold))
        self.btn_generar.setStyleSheet("""
            QPushButton {
                background-color: #0066B3;
                color: white;
                padding: 20px;
                border: none;
                border-radius: 10px;
            }
            QPushButton:hover {
                background-color: #0052A3;
            }
            QPushButton:disabled {
                background-color: #cccccc;
                color: #666666;
            }
        """)
        self.btn_generar.clicked.connect(self.generar_pdf)
        self.btn_generar.setEnabled(False)
        main_layout.addWidget(self.btn_generar)
        
        # Maximizar ventana al final (después de crear toda la interfaz)
        self.showMaximized()
    
    def cargar_productos(self):
        """Cargar productos desde el Excel"""
        try:
            if not os.path.exists(EXCEL_PATH):
                QMessageBox.critical(self, "Error", 
                    f"No se encontró el archivo:\n{EXCEL_PATH}\n\n"
                    "Asegúrate de que el archivo Excel esté en la carpeta correcta.")
                return
            
            wb = openpyxl.load_workbook(EXCEL_PATH)
            sheet = wb["Productos_FDA"]
            
            # Leer encabezados
            headers = []
            for col in range(1, 50):
                cell_value = sheet.cell(row=1, column=col).value
                if cell_value:
                    headers.append(cell_value)
                elif col > 10 and not cell_value:
                    break
            
            # Leer productos
            self.productos = []
            row = 2
            while True:
                nombre = sheet.cell(row=row, column=1).value
                if not nombre or nombre.strip() == "":
                    break
                
                producto = {}
                for col_index, header in enumerate(headers, 1):
                    valor = sheet.cell(row=row, column=col_index).value
                    producto[header] = str(valor) if valor is not None else ""
                
                self.productos.append(producto)
                row += 1
            
            wb.close()
            
            # Mostrar productos en la interfaz
            self.mostrar_productos()
            
            # Mensaje eliminado para no retrasar el flujo
        
        except Exception as e:
            QMessageBox.critical(self, "Error", 
                f"Error al cargar el Excel:\n{str(e)}")
    
    def mostrar_productos(self):
        """Mostrar productos en la interfaz"""
        # Limpiar layout anterior correctamente
        for i in reversed(range(self.productos_layout.count())): 
            widget = self.productos_layout.itemAt(i).widget()
            if widget is not None:
                widget.deleteLater()
        
        self.checkboxes = []
        self.spinboxes = []
        
        # Obtener filtro de búsqueda (si existe)
        filtro = ""
        if hasattr(self, 'search_input'):
            filtro = self.search_input.text().lower().strip()
        
        # Filtrar productos
        productos_filtrados = []
        for i, producto in enumerate(self.productos):
            nombre = str(producto.get('Product_Name', '')).lower().strip()
            
            # Si no hay filtro, incluir todos
            if not filtro:
                productos_filtrados.append((i, producto))
            # Si hay filtro, solo incluir si coincide
            elif filtro in nombre:
                productos_filtrados.append((i, producto))
        
        # Si no hay resultados, mostrar mensaje
        if len(productos_filtrados) == 0:
            no_results = QLabel("📦 No se encontraron productos con ese nombre")
            no_results.setFont(QFont("Arial", 14))
            no_results.setStyleSheet("color: #6c757d; padding: 40px;")
            no_results.setAlignment(Qt.AlignCenter)
            self.productos_layout.addWidget(no_results)
            self.productos_layout.addStretch()
            return
        
        for idx_original, producto in productos_filtrados:
            # Frame para cada producto
            frame = QFrame()
            frame.setStyleSheet("""
                QFrame {
                    background-color: #ffffff;
                    border: 2px solid #dee2e6;
                    border-radius: 8px;
                    padding: 10px;
                    margin: 5px;
                }
                QFrame:hover {
                    border-color: #0066B3;
                }
            """)
            
            layout = QHBoxLayout()
            frame.setLayout(layout)
            
            # Checkbox
            checkbox = QCheckBox()
            checkbox.setStyleSheet("QCheckBox::indicator { width: 25px; height: 25px; }")
            checkbox.setProperty("producto_index", idx_original)  # Guardar índice original
            checkbox.stateChanged.connect(self.actualizar_resumen)
            self.checkboxes.append(checkbox)
            layout.addWidget(checkbox)
            
            # Nombre del producto
            nombre_label = QLabel(producto.get('Product_Name', 'Sin nombre'))
            nombre_label.setFont(QFont("Arial", 11, QFont.Bold))
            nombre_label.setStyleSheet("color: #2c3e50;")  # Color explícito para Mac
            layout.addWidget(nombre_label, stretch=1)
            
            # Detalles
            detalles = f"Porción: {producto.get('Serving_Size', '')} | " \
                      f"Calorías: {producto.get('Calories', '')} | " \
                      f"Peso: {producto.get('Net_Weight', '')} | " \
                      f"Vence: {producto.get('Expiration_Date', '')}"
            detalles_label = QLabel(detalles)
            detalles_label.setFont(QFont("Arial", 9))
            detalles_label.setStyleSheet("color: #6c757d;")  # Color explícito para Mac
            layout.addWidget(detalles_label, stretch=2)
            
            # Cantidad
            cantidad_label = QLabel("Cantidad:")
            cantidad_label.setFont(QFont("Arial", 10, QFont.Bold))
            cantidad_label.setStyleSheet("color: #2c3e50;")  # Color explícito para Mac
            layout.addWidget(cantidad_label)
            
            spinbox = QSpinBox()
            spinbox.setMinimum(1)
            spinbox.setMaximum(999)
            spinbox.setValue(1)
            spinbox.setFont(QFont("Arial", 12, QFont.Bold))
            spinbox.setStyleSheet("""
                QSpinBox {
                    padding: 5px;
                    border: 2px solid #0066B3;
                    border-radius: 5px;
                    min-width: 80px;
                }
            """)
            spinbox.setProperty("producto_index", idx_original)  # Guardar índice original
            spinbox.valueChanged.connect(self.actualizar_resumen)
            self.spinboxes.append(spinbox)
            layout.addWidget(spinbox)
            
            self.productos_layout.addWidget(frame)
        
        self.productos_layout.addStretch()
    
    def filtrar_productos(self):
        """Filtrar productos según el texto de búsqueda"""
        self.mostrar_productos()
    
    def seleccionar_todo(self):
        """Seleccionar todos los productos"""
        for checkbox in self.checkboxes:
            checkbox.setChecked(True)
    
    def deseleccionar_todo(self):
        """Deseleccionar todos los productos"""
        for checkbox in self.checkboxes:
            checkbox.setChecked(False)
    
    def actualizar_resumen(self):
        """Actualizar el resumen de productos/etiquetas/hojas"""
        productos_sel = 0
        total_etiquetas = 0
        
        for i, checkbox in enumerate(self.checkboxes):
            if checkbox.isChecked():
                productos_sel += 1
                total_etiquetas += self.spinboxes[i].value()
        
        # Calcular hojas según tipo de etiqueta
        tipo_index = self.tipo_combo.currentIndex()
        
        # Si no ha seleccionado tipo (índice 0), no calcular hojas
        if tipo_index == 0:
            hojas = 0
        elif tipo_index == 4:  # PLS504
            hojas = (total_etiquetas + 9) // 10
        elif tipo_index == 5:  # Dulces - Avery 5260 (30 por hoja)
            hojas = (total_etiquetas + 29) // 30
        else:  # AVERY_8164, LACTEO_AVERY_8164, o LACTEO_SAN_JULIAN (todos 6 por hoja)
            hojas = (total_etiquetas + 5) // 6
        
        self.lbl_productos_sel.setText(str(productos_sel))
        self.lbl_total_etiquetas.setText(str(total_etiquetas))
        self.lbl_hojas.setText(str(hojas) if tipo_index > 0 else "—")
        
        # Habilitar/deshabilitar botón
        # Solo se habilita si: hay productos seleccionados Y se seleccionó tipo de etiqueta
        puede_generar = productos_sel > 0 and tipo_index > 0
        self.btn_generar.setEnabled(puede_generar)
    
    def generar_pdf(self):
        """Generar el PDF con las etiquetas seleccionadas"""
        try:
            # VALIDACIÓN 1: Verificar que se haya seleccionado tipo de etiqueta
            tipo_index = self.tipo_combo.currentIndex()
            if tipo_index == 0:
                QMessageBox.warning(self, "Tipo de Etiqueta Requerido", 
                    "⚠️ Debes seleccionar un tipo de etiqueta antes de generar el PDF.\n\n"
                    "Por favor, elige:\n"
                    "• Avery 8164 (vertical)\n"
                    "• Lacteo Avery 8164 (horizontal)\n"
                    "• Lacteo San Julian (vertical rotada)\n"
                    "• PLS 504 (pequeña)\n"
                    "• Dulces - Avery 5260")
                return
            
            # Obtener productos seleccionados
            productos_seleccionados = []
            for i, checkbox in enumerate(self.checkboxes):
                if checkbox.isChecked():
                    # Obtener índice original del producto
                    idx_original = checkbox.property("producto_index")
                    cantidad = self.spinboxes[i].value()
                    # Duplicar producto según cantidad
                    for _ in range(cantidad):
                        productos_seleccionados.append(self.productos[idx_original])
            
            if not productos_seleccionados:
                QMessageBox.warning(self, "Advertencia", 
                    "Debes seleccionar al menos un producto.")
                return
            
            # Determinar tipo de etiqueta (restar 1 porque índice 0 es el placeholder)
            if tipo_index == 1:
                label_type = 'AVERY_8164'
            elif tipo_index == 2:
                label_type = 'LACTEO_AVERY_8164'
            elif tipo_index == 3:
                label_type = 'LACTEO_SAN_JULIAN'
            elif tipo_index == 4:
                label_type = 'PLS504'
            else:  # tipo_index == 5
                label_type = 'AVERY_5260'
            
            cfg = LABEL_CONFIGS[label_type]
            
            # Generar nombre de archivo con timestamp
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            pdf_filename = os.path.join(BASE_DIR, f"Etiquetas_{timestamp}.pdf")
            
            # Crear PDF
            c = canvas.Canvas(pdf_filename, pagesize=letter, 
                            invariant=0, pageCompression=0)
            
            # Metadata
            c.setAuthor(f"Amaya_{timestamp}")
            c.setTitle(f"Etiquetas_FDA_{timestamp}")
            
            # Calcular posiciones
            pos = []
            for f in range(cfg['rows']):
                for cl in range(cfg['columns']):
                    pos.append((
                        cfg['margin_left'] + cl * (cfg['width'] + cfg['h_spacing']),
                        PAGE_HEIGHT - cfg['margin_top'] - (f + 1) * cfg['height'] - f * cfg['v_spacing']
                    ))
            
            # Dibujar etiquetas
            total = len(productos_seleccionados)
            for i, p in enumerate(productos_seleccionados):
                if i % cfg['per_page'] == 0 and i > 0:
                    c.showPage()
                
                x, y = pos[i % cfg['per_page']]
                
                if label_type == 'LACTEO_AVERY_8164':
                    dibujar_lacteo_avery(c, x, y, p, cfg)
                elif label_type == 'LACTEO_SAN_JULIAN':
                    dibujar_lacteo_san_julian(c, x, y, p, cfg)
                elif label_type == 'PLS504':
                    dibujar_pls(c, x, y, p, cfg)
                elif label_type == 'AVERY_5260':
                    dibujar_avery5260(c, x, y, p, cfg)
                else:
                    dibujar_avery(c, x, y, p, cfg)
            
            c.save()
            
            # Abrir PDF según el sistema operativo
            if platform.system() == "Windows":
                os.startfile(pdf_filename)
            elif platform.system() == "Darwin":  # Mac
                os.system(f'open "{pdf_filename}"')
            else:  # Linux
                os.system(f'xdg-open "{pdf_filename}"')
            
            # Mensaje de éxito
            QMessageBox.information(self, "¡Éxito!", 
                f"PDF generado exitosamente:\n\n"
                f"📄 Archivo: {os.path.basename(pdf_filename)}\n"
                f"🏷️ Total de etiquetas: {total}\n"
                f"📋 Hojas: {self.lbl_hojas.text()}\n"
                f"🎯 Tipo: {label_type}")
        
        except Exception as e:
            QMessageBox.critical(self, "Error", 
                f"Error al generar PDF:\n{str(e)}")


# ============================================================================
# FUNCIÓN PRINCIPAL
# ============================================================================
def main():
    app = QApplication(sys.argv)
    
    # Verificar que exista el directorio
    if not os.path.exists(BASE_DIR):
        msg = QMessageBox()
        msg.setIcon(QMessageBox.Critical)
        msg.setWindowTitle("Error")
        msg.setText(f"No existe la carpeta:\n{BASE_DIR}\n\n"
                   "Por favor, crea la carpeta y coloca los archivos necesarios.")
        msg.exec_()
        sys.exit(1)
    
    ventana = EtiquetasApp()
    ventana.show()
    sys.exit(app.exec_())


if __name__ == "__main__":
    main()